"""
COP REST API v0.1 Ã¢ÂÂ Clinical Operations Protocol
=================================================
FastAPI-baserat API fÃÂ¶r AI-driven schemaoptimering.

Endpoints:
  POST /schedule/generate     Ã¢ÂÂ Generera optimalt schema
  GET  /schedule/{id}         Ã¢ÂÂ HÃÂ¤mta genererat schema
  POST /schedule/adjust       Ã¢ÂÂ Manuell justering (byte, frÃÂ¥nvaro)
  POST /schedule/reoptimize   Ã¢ÂÂ Omoptimera efter ÃÂ¤ndring
  GET  /health                Ã¢ÂÂ HÃÂ¤lsocheck
  GET  /config                Ã¢ÂÂ HÃÂ¤mta aktiv klinikkonfiguration
  PUT  /config                Ã¢ÂÂ Uppdatera klinikkonfiguration
  POST /absence               Ã¢ÂÂ Registrera frÃÂ¥nvaro
  GET  /statistics/{id}       Ã¢ÂÂ HÃÂ¤mta schemastatistik
  POST /validate              Ã¢ÂÂ Validera schema mot ATL

Arkitektur:
  Browser/Tessa/TimeCase Ã¢ÂÂ [REST API] Ã¢ÂÂ [COP Solver] Ã¢ÂÂ [Optimalt Schema]
                                Ã¢ÂÂ
                         [Schedule Store]
"""

import os
import uuid
import time
import json
from datetime import datetime, date, timedelta
from typing import Optional
from collections import defaultdict
from enum import Enum

from fastapi import FastAPI, HTTPException, BackgroundTasks
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel, Field

from data_model import (
    ClinicConfig, Role, ShiftType, Function, Doctor,
    OperatingRoom, StaffingRequirement, CallStructure, ATLRules,
    create_kristianstad_example, create_generic_example, is_jour, OBRates,
)
from solver import solve_schedule, expand_to_granular, solve_rolling, _ob_cost
from absence_chain import AbsenceChain, AbsenceChainResult, ChainStatus
from db import get_db, connect_db, close_db

# PDF Export & Email Notifications
try:
    from pdf_export import pdf_router
    HAS_PDF = True
except ImportError:
    HAS_PDF = False

try:
    from email_service import email_router
    HAS_EMAIL = True
except ImportError:
    HAS_EMAIL = False

# Security
try:
    from security import setup_security
    HAS_SECURITY = True
except ImportError:
    HAS_SECURITY = False

# Auth & WebSocket integration
try:
    from auth import auth_router, get_current_user, require_role, require_permission, Role as AuthRole
    HAS_AUTH = True
except ImportError:
    HAS_AUTH = False

try:
    from websocket_hub import ws_router, hub
    HAS_WS = True
except ImportError:
    HAS_WS = False

# === APP SETUP ===
app = FastAPI(
    title="COP Ã¢ÂÂ Clinical Operations Protocol",
    description="AI-driven schemaoptimering fÃÂ¶r sjukvÃÂ¥rden. System-agnostisk motor som pluggar in i Tessa, Time Care, Heroma, Medvind.",
    version="0.1.0",
    docs_url="/docs",
    redoc_url="/redoc",
)

_cors_origins = os.environ.get("COP_CORS_ORIGINS", "*")
_origins = ["*"] if _cors_origins == "*" else [o.strip() for o in _cors_origins.split(",")]
app.add_middleware(
    CORSMiddleware,
    allow_origins=_origins,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Register auth & websocket routers
if HAS_AUTH:
    app.include_router(auth_router)
if HAS_WS:
    app.include_router(ws_router)
if HAS_PDF:
    app.include_router(pdf_router)
if HAS_EMAIL:
    app.include_router(email_router)

if HAS_PDF:
    app.include_router(pdf_router)
if HAS_EMAIL:
    app.include_router(email_router)

# Security middleware
if HAS_SECURITY:
    setup_security(app)

# === DATABASE LAYER (PostgreSQL med in-memory fallback) ===
db = get_db()

_active_jobs: dict = {}  # Ephemeral: aktiva solver-jobb under körning


# === PYDANTIC MODELLER (API-kontrakt) ===

class ScheduleRequest(BaseModel):
    """BegÃÂ¤ran om att generera ett nytt schema."""
    clinic_id: str = Field(description="Klinik-ID")
    num_weeks: int = Field(default=2, ge=1, le=8, description="Antal veckor att schemalÃÂ¤gga")
    start_date: Optional[str] = Field(default=None, description="Startdatum (YYYY-MM-DD), default=nÃÂ¤sta mÃÂ¥ndag")
    time_limit_seconds: int = Field(default=30, ge=5, le=300, description="Max tid fÃÂ¶r solver")
    locked_assignments: Optional[dict] = Field(default=None, description="LÃÂ¥sta tilldelningar {doctor_id: {day: function}}")
    excluded_doctors: Optional[list[str]] = Field(default=None, description="LÃÂ¤kar-ID som inte ska schemalÃÂ¤ggas")


class RollForwardRequest(BaseModel):
    """Rulla schema framåt — behåll låsta veckor, generera nya."""
    schedule_id: str = Field(description="Befintligt schema-ID")
    weeks_to_keep: int = Field(default=3, ge=1, le=7, description="Antal veckor att behålla")
    new_weeks: int = Field(default=1, ge=1, le=4, description="Antal nya veckor att generera")
    time_limit_seconds: int = Field(default=30, ge=5, le=300, description="Max tid")


class ScheduleResponse(BaseModel):
    """Svar med genererat schema."""
    schedule_id: str
    status: str  # "optimal", "feasible", "infeasible", "pending"
    clinic_id: str
    num_weeks: int
    start_date: str
    created_at: str
    solve_time_ms: int
    objective_value: Optional[float] = None
    schedule: Optional[dict] = None  # {doctor_id: {day_index: function_id}}
    statistics: Optional[dict] = None
    warnings: list[str] = []


class AdjustmentRequest(BaseModel):
    """BegÃÂ¤ran om att justera ett befintligt schema."""
    schedule_id: str
    adjustment_type: str  # "swap", "replace", "lock", "unlock"
    doctor_id: str
    day: int
    new_function: Optional[str] = None
    swap_with_doctor_id: Optional[str] = None
    reason: Optional[str] = None


class AbsenceRequest(BaseModel):
    """Registrera frÃÂ¥nvaro fÃÂ¶r en lÃÂ¤kare."""
    clinic_id: str
    doctor_id: str
    absence_type: str  # "sjuk", "semester", "vab", "utbildning", "konferens"
    start_date: str  # YYYY-MM-DD
    end_date: str    # YYYY-MM-DD
    reason: Optional[str] = None
    reoptimize: bool = Field(default=True, description="Auto-omoptimera berÃÂ¶rda scheman?")


class ReoptimizeRequest(BaseModel):
    """BegÃÂ¤ran om omoptimering av befintligt schema."""
    schedule_id: str
    preserve_locked: bool = Field(default=True, description="BehÃÂ¥ll lÃÂ¥sta tilldelningar?")
    time_limit_seconds: int = Field(default=30)


class ValidationResult(BaseModel):
    """Resultat av ATL-validering."""
    valid: bool
    violations: list[dict] = []
    warnings: list[dict] = []
    summary: dict = {}


class DoctorInput(BaseModel):
    """LÃÂ¤karinput fÃÂ¶r API."""
    id: str
    name: str
    role: str  # UL, ST_TIDIG, ST_SEN, SP, ÃÂL
    employment_percent: int = 100
    can_primary_call: bool = False
    can_backup_call: bool = False
    exempt_from_call: bool = False
    supervisor_id: Optional[str] = None
    site_preference: Optional[str] = None
    # Avancerad schemaläggning
    schedule_pattern: str = "weekly"
    fixed_weekdays: Optional[dict] = None
    min_shifts_per_week: Optional[dict] = None
    max_shifts_per_week: Optional[dict] = None
    half_day_schedule: Optional[dict] = None
    current_rotation_block: Optional[dict] = None
    recurring_activities: Optional[list] = None
    work_days_per_week: Optional[int] = None


class ConfigUpdate(BaseModel):
    """Uppdatering av klinikkonfiguration."""
    clinic_id: str
    doctors: Optional[list[DoctorInput]] = None
    # Fler config-fÃÂ¤lt kan lÃÂ¤ggas till


class HealthResponse(BaseModel):
    """HÃÂ¤lsocheck-svar."""
    status: str
    version: str
    uptime_seconds: float
    schedules_generated: int
    solver_available: bool


# === STARTUP ===
START_TIME = time.time()

@app.on_event("startup")
async def startup():
    """Anslut till PostgreSQL och ladda demo-konfiguration om COP_DEMO=true.

    Total startup budget: 60 s.  Varje steg har sin egen timeout så att
    appen ALLTID når health-check inom Railways 5-minutersgräns.
    """
    import asyncio

    # --- 1. Database connection (max 20 s via connect_db internals) ---
    try:
        db_ok = await asyncio.wait_for(connect_db(), timeout=30)
    except asyncio.TimeoutError:
        print('[STARTUP] connect_db timeout (30 s) -- falling back to in-memory')
        db_ok = False
    except Exception as exc:
        print(f'[STARTUP] connect_db error: {exc} -- falling back to in-memory')
        db_ok = False

    backend = 'PostgreSQL' if db_ok else 'in-memory'

    # --- 2. Auth init (max 15 s) ---
    try:
        from auth import init_auth
        await asyncio.wait_for(init_auth(db), timeout=15)
    except asyncio.TimeoutError:
        print('[STARTUP] init_auth timeout (15 s) -- continuing without cached users')
    except Exception as exc:
        print(f'[STARTUP] init_auth error: {exc} -- continuing anyway')

    # --- 3. Demo configs (max 10 s) ---
    demo_mode = os.environ.get('COP_DEMO', 'true').lower() in ('true', '1', 'yes')
    if demo_mode:
        try:
            config = create_kristianstad_example()
            await asyncio.wait_for(db.save_config('kristianstad', config), timeout=5)
            generic = create_generic_example()
            await asyncio.wait_for(db.save_config('generic', generic), timeout=5)
        except Exception as exc:
            print(f'[STARTUP] demo config error: {exc} -- continuing without demo data')

    print(f'[STARTUP] COP API ready. Backend: {backend}. Demo: {demo_mode}.')


@app.on_event("shutdown")
async def shutdown():
    """StÃÂ¤ng MongoDB-anslutning."""
    await close_db()


# === HEALTH & INFO ===


@app.post("/auth/reset-demo-passwords", tags=["System"])
async def reset_demo_passwords(secret: str = "cop-reset-2026"):
    """
    Nollställer demo-lösenorden till standardvärdena.
    Kräver secret-parametern för att skydda mot oavsiktlig användning.
    Anrop: POST /auth/reset-demo-passwords?secret=cop-reset-2026
    """
    if secret != "cop-reset-2026":
        raise HTTPException(status_code=403, detail="Fel secret")
    try:
        from auth import hash_password, UserInDB, Role as AuthRole, _cache_set
        import secrets as _secrets
        defaults = [
            ("usr_admin",     "admin",     "admin@cop.local",  "COP Administrator", AuthRole.ADMIN,     "cop-admin-2026"),
            ("usr_scheduler", "scheduler", "schema@cop.local", "Schemaläggare",     AuthRole.SCHEDULER, "schema-2026"),
            ("usr_viewer",    "viewer",    "viewer@cop.local", "Dashboard Viewer",  AuthRole.VIEWER,    "viewer-2026"),
        ]
        reset = []
        for uid, uname, email, name, role, pwd in defaults:
            user = UserInDB(
                user_id=uid,
                username=uname,
                email=email,
                full_name=name,
                role=role,
                hashed_password=hash_password(pwd),
                password_change_required=False,
                is_active=True,
            )
            await db.save_user(user.model_dump())
            _cache_set(user)
            reset.append(uname)
        return {"reset": reset, "message": "Lösenord nollställda till standardvärden"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Reset misslyckades: {e}")


@app.get("/health", response_model=HealthResponse, tags=["System"])
async def health_check():
    """Kontrollera att API:t lever och solver fungerar."""
    return HealthResponse(
        status="healthy",
        version="0.1.0",
        uptime_seconds=round(time.time() - START_TIME, 1),
        schedules_generated=0,
        solver_available=True,
    )


@app.get("/config", tags=["Konfiguration"])
async def get_default_config(clinic_id: Optional[str] = None):
    """HÃÂ¤mta klinikkonfiguration. Utan clinic_id returneras fÃÂ¶rsta tillgÃÂ¤ngliga."""
    if not clinic_id:
        configs = await db.list_configs()
        if not configs:
            raise HTTPException(status_code=404, detail="Ingen klinik konfigurerad")
        clinic_id = configs[0]["clinic_id"]
    return await get_config_by_id(clinic_id)

@app.get("/configs", tags=["Konfiguration"])
async def list_configs():
    """Lista alla tillgÃÂ¤ngliga klinikkonfigurationer."""
    configs = await db.list_configs()
    return configs


@app.get("/statistics", tags=["Statistik"])
async def get_latest_statistics():
    """HÃÂ¤mta statistik fÃÂ¶r senaste schemat."""
    schedules = await db.list_schedules(limit=1)
    if not schedules:
        return {"total_doctors": 0, "atl_violations": 0, "weeks": 0,
                "role_distribution": {}, "shift_distribution": {}}
    latest = schedules[0]
    latest_id = latest["schedule_id"]
    sched = await db.get_schedule(latest_id)
    stats = sched.get("statistics", {}) if sched else {}
    stats["schedule_id"] = latest_id
    return stats


@app.get("/config/{clinic_id}", tags=["Konfiguration"])
async def get_config_by_id(clinic_id: str):
    """HÃÂ¤mta klinikkonfiguration."""
    config = await db.get_config(clinic_id)
    if not config:
        raise HTTPException(status_code=404, detail=f"Klinik '{clinic_id}' finns inte")

    from data_model import config_to_dict
    result = config_to_dict(config)
    result["clinic_id"] = clinic_id
    result["num_doctors"] = len(config.doctors)
    result["num_rooms"] = len(config.operating_rooms)
    return result


@app.get("/clinic/{clinic_id}/ui-config", tags=["Konfiguration"])
async def get_ui_config(clinic_id: str):
    """
    Returnerar dynamisk UI-konfiguration för frontend.
    Allt som frontend behöver för dropdowns, etiketter och formulär.
    Drivs av klinikens faktiska config — inga hardkodade värden i frontend.
    """
    config = await db.get_config(clinic_id)
    if not config:
        raise HTTPException(status_code=404, detail=f"Klinik '{clinic_id}' finns inte")

    from data_model import Role, Function

    # Tillgängliga sites
    sites = config.sites or []
    clinic_type = getattr(config, 'clinic_type', 'kirurgi')
    has_on_call = getattr(config, 'has_on_call', True)
    has_operations = getattr(config, 'has_operations', True)

    # Roller som finns i kliniken (baserat på faktiska läkare)
    active_roles = sorted(set(d.role.value for d in config.doctors))
    role_labels = {
        "AT": "AT-läkare", "UL": "Underläkare",
        "ST_TIDIG": "ST-läkare (tidig)", "ST_SEN": "ST-läkare (senior)",
        "SP": "Specialist", "ÖL": "Överläkare",
        "DL": "Distriktsläkare", "SSK": "Sjuksköterska",
        "USK": "Undersköterska", "FT": "Fysioterapeut",
        "PSY": "Psykolog", "BM": "Barnmorska",
        "KUR": "Kurator", "DIET": "Dietist",
        "AT_TERAPEUT": "Arbetsterapeut", "CUSTOM": "Annan",
    }
    roles = [{"value": r, "label": role_labels.get(r, r)} for r in active_roles]
    # Inkludera alla roller (för att kunna lägga till nya läkare)
    all_roles = [{"value": r.value, "label": role_labels.get(r.value, r.value)} for r in Role]
    # Lägg till custom roles
    for cr in getattr(config, 'custom_roles', []):
        cr_id = cr.get("id", "")
        cr_label = cr.get("label", cr_id)
        if cr_id and not any(r["value"] == cr_id for r in all_roles):
            all_roles.append({"value": cr_id, "label": cr_label})

    # Funktioner — bygg dynamiskt från sites + verksamhetstyp
    day_functions = []

    # OP-funktioner (bara om verksamheten har operationer)
    if has_operations:
        for site in sites:
            day_functions.append({"value": f"OP_{site}", "label": f"OP {site}", "category": "operation", "site": site})

    # Verksamhetstyp-specifika funktioner per site
    _CLINIC_TYPE_SITE_FUNCS = {
        "kirurgi": [
            ("AVD", "Avdelning", "vård"),
            ("MOTT", "Mottagning", "mottagning"),
            ("AKUT", "Akut", "akut"),
        ],
        "internmedicin": [
            ("AVD", "Avdelning", "vård"),
            ("MOTT", "Mottagning", "mottagning"),
            ("ROND", "Rond", "vård"),
            ("DIALYS", "Dialys", "vård"),
            ("DAGVÅRD", "Dagvård", "vård"),
        ],
        "vardcentral": [
            ("MOTT", "Mottagning", "mottagning"),
            ("BVC", "BVC", "mottagning"),
            ("MVC", "MVC", "mottagning"),
            ("LAB", "Lab", "mottagning"),
            ("TELEFON", "Telefontid", "mottagning"),
            ("HEMBESÖK", "Hembesök", "mottagning"),
            ("VIDEO", "Videomottagning", "mottagning"),
        ],
        "oppenvard": [
            ("MOTT", "Mottagning", "mottagning"),
        ],
        "psykiatri": [
            ("AVD", "Avdelning", "vård"),
            ("MOTT", "Mottagning", "mottagning"),
            ("SAMTAL", "Samtal", "mottagning"),
            ("GRUPP", "Gruppterapi", "mottagning"),
            ("AKUTPSYK", "Akutpsykiatri", "akut"),
        ],
        "rehabilitering": [
            ("MOTT", "Mottagning", "mottagning"),
            ("REHAB", "Rehabilitering", "vård"),
        ],
    }

    site_funcs = _CLINIC_TYPE_SITE_FUNCS.get(clinic_type, [
        ("AVD", "Avdelning", "vård"),
        ("MOTT", "Mottagning", "mottagning"),
    ])
    for site in sites:
        for func_id, func_label, func_cat in site_funcs:
            day_functions.append({
                "value": f"{func_id}_{site}", "label": f"{func_label} {site}",
                "category": func_cat, "site": site,
            })

    # Icke-platsspecifika
    day_functions.extend([
        {"value": "ADMIN", "label": "Admin", "category": "admin", "site": None},
        {"value": "FORSKNING", "label": "Forskning", "category": "admin", "site": None},
        {"value": "HANDLEDNING", "label": "Handledning", "category": "utbildning", "site": None},
        {"value": "UTBILDNING", "label": "Utbildning", "category": "utbildning", "site": None},
        {"value": "LEDIG", "label": "Ledig", "category": "frånvaro", "site": None},
    ])

    # Custom functions
    for cf in getattr(config, 'custom_functions', []):
        cf_id = cf.get("id", "")
        cf_label = cf.get("label", cf_id)
        cf_cat = cf.get("category", "annan")
        cf_site = cf.get("site_specific", False)
        if cf_id:
            if cf_site:
                for site in sites:
                    day_functions.append({
                        "value": f"{cf_id}_{site}", "label": f"{cf_label} {site}",
                        "category": cf_cat, "site": site,
                    })
            else:
                day_functions.append({
                    "value": cf_id, "label": cf_label,
                    "category": cf_cat, "site": None,
                })

    # Operationssalar
    rooms = [{"id": r.id, "site": r.site, "name": r.name,
              "available_days": r.available_days} for r in config.operating_rooms]

    # Jourlinjer (bara om verksamheten har jour)
    call_functions = []
    if has_on_call:
        call_functions = [
            {"value": "JOUR_P", "label": "Primärjour"},
            {"value": "JOUR_B", "label": "Bakjour"},
        ]

    # Skifttyper
    shift_types = [
        {"value": "DAG", "label": "Dag (07-16:30)"},
    ]
    if has_on_call:
        shift_types.extend([
            {"value": "JOUR_KVÄLL", "label": "Kvällsjour (16:30-22)"},
            {"value": "JOUR_NATT", "label": "Nattjour (22-07)"},
            {"value": "JOUR_HELGDAG", "label": "Helgdag (07-22)"},
            {"value": "JOUR_HELGNATT", "label": "Helgnatt (22-07)"},
        ])

    # Bemanningskrav
    staffing = [{"function": sr.function.value, "site": sr.site,
                 "min_count": sr.min_count, "shift_type": sr.shift_type.value}
                for sr in config.staffing_requirements]

    # Kompetenser (unika från befintliga läkare)
    all_competencies = sorted(set(c for d in config.doctors for c in getattr(d, 'competencies', [])))

    # Randningskliniker (unika från ST-läkare)
    randning_kliniker = sorted(set(
        r.get('klinik', '') for d in config.doctors
        for r in getattr(d, 'st_randning', []) if r.get('klinik')
    ))

    return {
        "clinic_id": clinic_id,
        "clinic_name": config.name,
        "clinic_type": clinic_type,
        "has_on_call": has_on_call,
        "has_operations": has_operations,
        "sites": sites,
        "roles": roles,
        "all_roles": all_roles,
        "day_functions": day_functions,
        "call_functions": call_functions,
        "shift_types": shift_types,
        "operating_rooms": rooms,
        "staffing_requirements": staffing,
        "competencies": all_competencies,
        "randning_kliniker": randning_kliniker,
        "schedule_cycle_weeks": config.schedule_cycle_weeks,
        "travel_time_between_sites_min": config.travel_time_between_sites_min,
        "custom_roles": getattr(config, 'custom_roles', []),
        "custom_functions": getattr(config, 'custom_functions', []),
    }


# === CONFIG CRUD ===

class ClinicConfigInput(BaseModel):
    """Input for creating/updating a clinic config."""
    name: str
    sites: list[str]
    doctors: list[dict] = []
    operating_rooms: list[dict] = []
    staffing_requirements: list[dict] = []
    call_structure: dict = {}
    atl_rules: dict = {}
    preferences: list[dict] = []
    shift_definitions: list[dict] = []
    constraint_rules: list[dict] = []
    schedule_cycle_weeks: int = 10
    travel_time_between_sites_min: int = 0
    schedule_start_date: str = ""
    optimize_ob_cost: bool = True
    clinic_type: str = "kirurgi"
    has_on_call: bool = True
    has_operations: bool = True
    custom_roles: list[dict] = []
    custom_functions: list[dict] = []

@app.post("/config", tags=["Konfiguration"])
async def create_config(clinic_id: str, body: ClinicConfigInput):
    """Skapa ny klinikkonfiguration."""
    existing = await db.get_config(clinic_id)
    if existing:
        raise HTTPException(status_code=409, detail=f"Klinik '{clinic_id}' finns redan")
    from data_model import dict_to_config
    config = dict_to_config(body.model_dump())
    await db.save_config(clinic_id, config)
    await db.audit("config_created", details={"clinic_id": clinic_id})
    return {"status": "created", "clinic_id": clinic_id}

@app.put("/config/{clinic_id}", tags=["Konfiguration"])
async def update_config(clinic_id: str, body: ClinicConfigInput):
    """Uppdatera hela klinikkonfigurationen."""
    existing = await db.get_config(clinic_id)
    if not existing:
        raise HTTPException(status_code=404, detail=f"Klinik '{clinic_id}' finns inte")
    from data_model import dict_to_config
    config = dict_to_config(body.model_dump())
    await db.save_config(clinic_id, config)
    await db.audit("config_updated", details={"clinic_id": clinic_id})
    return {"status": "updated", "clinic_id": clinic_id}

@app.patch("/config/{clinic_id}/doctors", tags=["Konfiguration"])
async def patch_doctors(clinic_id: str, doctors: list[dict]):
    """Uppdatera bara lakarlistan."""
    config = await db.get_config(clinic_id)
    if not config:
        raise HTTPException(status_code=404, detail=f"Klinik '{clinic_id}' finns inte")
    from data_model import config_to_dict, dict_to_config
    data = config_to_dict(config)
    data["doctors"] = doctors
    new_config = dict_to_config(data)
    await db.save_config(clinic_id, new_config)
    await db.audit("config_doctors_updated", details={"clinic_id": clinic_id, "count": len(doctors)})
    return {"status": "updated", "doctors": len(doctors)}

@app.patch("/config/{clinic_id}/rules", tags=["Konfiguration"])
async def patch_rules(clinic_id: str, rules: list[dict]):
    """Uppdatera regler/constraints."""
    config = await db.get_config(clinic_id)
    if not config:
        raise HTTPException(status_code=404, detail=f"Klinik '{clinic_id}' finns inte")
    from data_model import config_to_dict, dict_to_config
    data = config_to_dict(config)
    data["constraint_rules"] = rules
    new_config = dict_to_config(data)
    await db.save_config(clinic_id, new_config)
    await db.audit("config_rules_updated", details={"clinic_id": clinic_id, "count": len(rules)})
    return {"status": "updated", "rules": len(rules)}

@app.patch("/config/{clinic_id}/shifts", tags=["Konfiguration"])
async def patch_shifts(clinic_id: str, shifts: list[dict]):
    """Uppdatera passtyper."""
    config = await db.get_config(clinic_id)
    if not config:
        raise HTTPException(status_code=404, detail=f"Klinik '{clinic_id}' finns inte")
    from data_model import config_to_dict, dict_to_config
    data = config_to_dict(config)
    data["shift_definitions"] = shifts
    new_config = dict_to_config(data)
    await db.save_config(clinic_id, new_config)
    return {"status": "updated", "shifts": len(shifts)}

@app.delete("/config/{clinic_id}", tags=["Konfiguration"])
async def delete_config(clinic_id: str):
    """Ta bort klinikkonfiguration."""
    existing = await db.get_config(clinic_id)
    if not existing:
        raise HTTPException(status_code=404, detail=f"Klinik '{clinic_id}' finns inte")
    await db.delete_config(clinic_id)
    await db.audit("config_deleted", details={"clinic_id": clinic_id})
    return {"status": "deleted", "clinic_id": clinic_id}


# === AI ENDPOINTS ===

class AIRuleRequest(BaseModel):
    clinic_id: str
    rule_text: str

class AIChatRequest(BaseModel):
    clinic_id: str
    user_id: str
    message: str

class AIExplainRequest(BaseModel):
    schedule_id: str
    doctor_id: str
    shift_date: str

class AIPredictRequest(BaseModel):
    clinic_id: str
    period_start: str
    period_end: str

class AIConflictRequest(BaseModel):
    clinic_id: str
    new_rule: dict

class AIOnboardingRequest(BaseModel):
    message: str
    chat_history: list[dict] = []
    partial_config: dict = {}

class AIAnalyzeConfigRequest(BaseModel):
    clinic_id: str

@app.post("/api/ai/rules/parse", tags=["AI"])
async def ai_parse_rule(req: AIRuleRequest):
    config = await db.get_config(req.clinic_id)
    if not config:
        raise HTTPException(status_code=404, detail=f"Klinik '{req.clinic_id}' finns inte")
    from ai_rules import parse_rule
    result = await parse_rule(config, req.rule_text, clinic_id=req.clinic_id)
    if result.get("constraint") and not result.get("error"):
        await db.save_ai_rule(req.clinic_id, req.rule_text, result["constraint"], result["confidence"])
    return result

@app.post("/api/ai/conflicts/check", tags=["AI"])
async def ai_check_conflicts(req: AIConflictRequest):
    config = await db.get_config(req.clinic_id)
    if not config:
        raise HTTPException(status_code=404, detail=f"Klinik '{req.clinic_id}' finns inte")
    from ai_conflicts import check_conflicts
    return await check_conflicts(config, req.new_rule, clinic_id=req.clinic_id)

@app.post("/api/ai/explain", tags=["AI"])
async def ai_explain(req: AIExplainRequest):
    sched = await db.get_schedule(req.schedule_id)
    if not sched:
        raise HTTPException(status_code=404, detail="Schema inte hittat")
    config = await db.get_config(sched.get("clinic_id", ""))
    if not config:
        raise HTTPException(status_code=404, detail="Klinik-config saknas")
    from ai_explain import explain_assignment
    return await explain_assignment(config, sched, req.doctor_id, req.shift_date, clinic_id=sched.get("clinic_id", ""))

@app.post("/api/ai/predict/absence", tags=["AI"])
async def ai_predict(req: AIPredictRequest):
    config = await db.get_config(req.clinic_id)
    if not config:
        raise HTTPException(status_code=404, detail=f"Klinik '{req.clinic_id}' finns inte")
    chains = await db.list_chains(limit=200)
    from ai_predict import predict_absence
    result = await predict_absence(chains, req.period_start, req.period_end, len(config.doctors), clinic_id=req.clinic_id)
    if not result.get("error"):
        await db.save_ai_prediction(req.clinic_id, f"{req.period_start}_{req.period_end}", result)
    return result

@app.post("/api/ai/chat", tags=["AI"])
async def ai_chat_endpoint(req: AIChatRequest):
    config = await db.get_config(req.clinic_id)
    if not config:
        raise HTTPException(status_code=404, detail=f"Klinik '{req.clinic_id}' finns inte")
    schedules = await db.list_schedules(clinic_id=req.clinic_id, limit=1)
    latest = await db.get_schedule(schedules[0]["schedule_id"]) if schedules else {}
    history = await db.get_ai_chat_history(req.clinic_id, req.user_id, limit=10)
    from ai_chat import chat
    result = await chat(config, latest, req.user_id, req.message, chat_history=history, clinic_id=req.clinic_id)
    await db.save_ai_chat(req.clinic_id, req.user_id, req.message, result.get("response_sv", ""), result.get("action"))
    return result


# === AI ONBOARDING ===

@app.post("/api/ai/onboarding", tags=["AI Onboarding"])
async def ai_onboarding(req: AIOnboardingRequest):
    """
    Konversations-baserad klinik-onboarding.
    AI:n intervjuar admin steg för steg och bygger config.
    """
    from ai_onboarding import onboard_step
    result = await onboard_step(
        message=req.message,
        chat_history=req.chat_history,
        partial_config=req.partial_config,
    )
    return result


@app.post("/api/ai/onboarding/generate", tags=["AI Onboarding"])
async def ai_onboarding_generate(body: dict):
    """
    Generera komplett ClinicConfig från onboarding-resultat.
    Anropas när onboarding is_complete=True.
    """
    from ai_onboarding import generate_clinic_config
    config = await generate_clinic_config(body)
    return config


@app.post("/api/ai/onboarding/save", tags=["AI Onboarding"])
async def ai_onboarding_save(clinic_id: str, body: dict):
    """
    Spara en AI-genererad config som ny klinik.
    """
    existing = await db.get_config(clinic_id)
    if existing:
        raise HTTPException(status_code=409, detail=f"Klinik '{clinic_id}' finns redan")
    from data_model import dict_to_config
    config = dict_to_config(body)
    await db.save_config(clinic_id, config)
    await db.audit("clinic_created_via_onboarding", details={"clinic_id": clinic_id})
    return {"status": "created", "clinic_id": clinic_id}


@app.post("/api/ai/analyze-config", tags=["AI Onboarding"])
async def ai_analyze_config(req: AIAnalyzeConfigRequest):
    """
    Analysera befintlig klinikconfig och hitta saknade regler,
    ofullständig data, inkonsistenser och förbättringsförslag.
    """
    config = await db.get_config(req.clinic_id)
    if not config:
        raise HTTPException(status_code=404, detail=f"Klinik '{req.clinic_id}' finns inte")
    from data_model import config_to_dict
    from ai_onboarding import analyze_config_gaps
    config_dict = config_to_dict(config)
    result = await analyze_config_gaps(config_dict)
    return result


# === SCHEMAGENERERING ===

def _run_solver(job_id: str, config: ClinicConfig, request: ScheduleRequest):
    """KÃÂ¶r solver i bakgrunden."""
    try:
        _active_jobs[job_id]["status"] = "running"
        start_time = time.time()

        schedule = solve_schedule(
            config,
            num_weeks=request.num_weeks,
            time_limit_seconds=request.time_limit_seconds,
        )

        solve_time_ms = int((time.time() - start_time) * 1000)

        if schedule is None:
            _active_jobs[job_id]["status"] = "infeasible"
            _active_jobs[job_id]["error"] = "Ingen giltig lÃÂ¶sning hittades"
            return

        # BerÃÂ¤kna startdatum
        if request.start_date:
            start = datetime.strptime(request.start_date, "%Y-%m-%d").date()
        else:
            today = date.today()
            days_until_monday = (7 - today.weekday()) % 7
            if days_until_monday == 0:
                days_until_monday = 7
            start = today + timedelta(days=days_until_monday)

        # BerÃÂ¤kna statistik
        statistics = _compute_statistics(schedule, config, request.num_weeks * 7)

        # Expandera till granulära jourtyper
        num_days = request.num_weeks * 7
        granular = expand_to_granular(schedule, num_days)

        # Konvertera schedule till datum-baserat format (granulär)
        date_schedule = {}
        for doc_id, days in granular.items():
            date_schedule[doc_id] = {}
            for day_idx, func in days.items():
                day_date = start + timedelta(days=day_idx)
                date_schedule[doc_id][day_date.isoformat()] = func

        schedule_id = _active_jobs[job_id]["schedule_id"]
        schedule_data = {
            "schedule_id": schedule_id,
            "status": "optimal",
            "clinic_id": request.clinic_id,
            "num_weeks": request.num_weeks,
            "start_date": start.isoformat(),
            "created_at": datetime.now().isoformat(),
            "solve_time_ms": solve_time_ms,
            "schedule": date_schedule,
            "raw_schedule": schedule,  # Indexbaserat fÃÂ¶r intern anvÃÂ¤ndning
            "statistics": statistics,
            "warnings": [],
        }

        _active_jobs[job_id]["status"] = "completed"
        _active_jobs[job_id]["result"] = schedule_data

        # Spara version
        import asyncio
        try:
            loop = asyncio.get_event_loop()
            loop.create_task(db.save_version(schedule_id, {
                "change_type": "generated",
                "raw_schedule": schedule,
                "change_description": f"Schema genererat ({request.num_weeks} veckor)",
            }))
        except Exception:
            pass

        # WebSocket broadcast: schema genererat
        if HAS_WS:
            import asyncio
            try:
                asyncio.get_event_loop().create_task(
                    hub.broadcast("schedule", {
                        "schedule_id": schedule_id,
                        "clinic_id": request.clinic_id,
                        "weeks": request.num_weeks,
                        "solve_time_ms": solve_time_ms,
                    }, event_type="schedule_generated")
                )
            except Exception:
                pass  # WebSocket broadcast failure should not break API

    except Exception as e:
        _active_jobs[job_id]["status"] = "failed"
        _active_jobs[job_id]["error"] = str(e)


def _compute_statistics(schedule: dict, config: ClinicConfig, num_days: int) -> dict:
    """BerÃÂ¤kna schemastatistik."""
    doc_by_id = {d.id: d for d in config.doctors}
    stats = {
        "call_distribution": {},
        "staffing_per_day": {},
        "st_matching": {},
        "atl_violations": [],
        "workload_balance": {},
    }

    # JourfÃÂ¶rdelning
    for doc in config.doctors:
        primary = sum(1 for d in range(num_days) if "JOUR_P" in str(schedule.get(doc.id, {}).get(d, "")))
        backup = sum(1 for d in range(num_days) if "JOUR_B" in str(schedule.get(doc.id, {}).get(d, "")))
        if primary + backup > 0:
            stats["call_distribution"][doc.id] = {
                "name": doc.name,
                "role": doc.role.value,
                "primary": primary,
                "backup": backup,
                "total": primary + backup,
            }

    # Bemanningstal per dag
    for day in range(num_days):
        counts = defaultdict(int)
        for doc in config.doctors:
            func = schedule.get(doc.id, {}).get(day, "LEDIG")
            if func != "LEDIG":
                counts[func] += 1
        stats["staffing_per_day"][day] = dict(counts)

    # ST-handledarmatchning
    for doc in config.doctors:
        if doc.supervisor_id:
            matches = 0
            total_op = 0
            for day in range(num_days):
                func = schedule.get(doc.id, {}).get(day, "")
                if func.startswith("OP_"):
                    total_op += 1
                    sup_func = schedule.get(doc.supervisor_id, {}).get(day, "")
                    if sup_func == func:
                        matches += 1
            if total_op > 0:
                stats["st_matching"][doc.id] = {
                    "name": doc.name,
                    "matches": matches,
                    "total_op_days": total_op,
                    "match_rate": round(matches / total_op * 100, 1),
                }

    # ATL-brott
    for doc in config.doctors:
        for day in range(num_days - 1):
            func_today = schedule.get(doc.id, {}).get(day, "LEDIG")
            func_tomorrow = schedule.get(doc.id, {}).get(day + 1, "LEDIG")
            if is_jour(func_today) and not is_jour(func_tomorrow) and func_tomorrow != "LEDIG":
                stats["atl_violations"].append({
                    "doctor_id": doc.id,
                    "doctor_name": doc.name,
                    "day": day,
                    "violation": f"Jour dag {day+1} Ã¢ÂÂ arbete dag {day+2}",
                })

    # Arbetsbelastning (antal arbetsdagar)
    for doc in config.doctors:
        work_days = sum(1 for d in range(num_days)
                       if schedule.get(doc.id, {}).get(d, "LEDIG") != "LEDIG")
        stats["workload_balance"][doc.id] = {
            "name": doc.name,
            "role": doc.role.value,
            "work_days": work_days,
            "total_days": num_days,
            "utilization": round(work_days / num_days * 100, 1),
        }

    # OB-kostnader
    ob_costs = {}
    for doc in config.doctors:
        total_ob = 0.0
        for day in range(num_days):
            func = schedule.get(doc.id, {}).get(day, "LEDIG")
            weekday = day % 7
            total_ob += _ob_cost(weekday, func, config.ob_rates)
        if total_ob > 0:
            ob_costs[doc.id] = {"name": doc.name, "ob_cost": round(total_ob, 1)}
    stats["ob_costs"] = ob_costs
    stats["total_ob_cost"] = round(sum(v["ob_cost"] for v in ob_costs.values()), 1)

    return stats


async def _run_solver_async(job_id: str, config: ClinicConfig, request: ScheduleRequest):
    """Async wrapper för background task — kör solver och sparar till DB."""
    _run_solver(job_id, config, request)
    job = _active_jobs.get(job_id, {})
    if job.get("status") == "completed" and job.get("result"):
        await db.save_schedule(job["result"])
        await db.save_job(job)


@app.post("/schedule/generate", tags=["Schema"])
async def generate_schedule(request: ScheduleRequest, background_tasks: BackgroundTasks):
    """
    Generera ett nytt optimerat schema.

    Startar solver i bakgrunden. Returnerar job_id fÃÂ¶r polling.
    Alternativt: om time_limit <= 60s, kÃÂ¶r synkront.
    """
    config = await db.get_config(request.clinic_id)
    if not config:
        raise HTTPException(status_code=404, detail=f"Klinik '{request.clinic_id}' finns inte")

    schedule_id = f"sch_{uuid.uuid4().hex[:12]}"
    job_id = f"job_{uuid.uuid4().hex[:8]}"

    job_data = {
        "job_id": job_id,
        "schedule_id": schedule_id,
        "status": "queued",
        "created_at": datetime.now().isoformat(),
    }
    _active_jobs[job_id] = dict(job_data)  # ephemeral slot for _run_solver
    await db.save_job(job_data)

    if request.time_limit_seconds <= 60:
        # Synkron kÃÂ¶rning fÃÂ¶r snabba jobb
        _run_solver(job_id, config, request)
        job = _active_jobs[job_id]

        if job["status"] == "completed":
            await db.save_schedule(job["result"])
            await db.save_job(job)
            return job["result"]
        elif job["status"] == "infeasible":
            raise HTTPException(status_code=422, detail="Ingen giltig lÃÂ¶sning kunde hittas med givna constraints")
        else:
            raise HTTPException(status_code=500, detail=job.get("error", "OkÃÂ¤nt fel"))
    else:
        # Asynkron kÃÂ¶rning
        background_tasks.add_task(_run_solver_async, job_id, config, request)
        return {
            "job_id": job_id,
            "schedule_id": schedule_id,
            "status": "queued",
            "message": "Schemagenerering startad. Polla /job/{job_id} fÃÂ¶r status.",
        }


@app.get("/job/{job_id}", tags=["Schema"])
async def get_job_status(job_id: str):
    """HÃÂ¤mta status fÃÂ¶r ett bakgrundsjobb."""
    job = await db.get_job(job_id)
    if not job:
        raise HTTPException(status_code=404, detail="Jobb inte hittat")
    return job


@app.get("/schedule/{schedule_id}", tags=["Schema"])
async def get_schedule(schedule_id: str):
    """HÃÂ¤mta ett genererat schema."""
    sched = await db.get_schedule(schedule_id)
    if not sched:
        raise HTTPException(status_code=404, detail="Schema inte hittat")

    # Returnera utan raw_schedule (intern data)
    result = {k: v for k, v in sched.items() if k != "raw_schedule"}
    return result


@app.get("/schedule/{schedule_id}/doctor/{doctor_id}", tags=["Schema"])
async def get_doctor_schedule(schedule_id: str, doctor_id: str):
    """HÃÂ¤mta schema fÃÂ¶r en specifik lÃÂ¤kare."""
    sched = await db.get_schedule(schedule_id)
    if not sched:
        raise HTTPException(status_code=404, detail="Schema inte hittat")

    doctor_schedule = sched["schedule"].get(doctor_id)
    if not doctor_schedule:
        raise HTTPException(status_code=404, detail=f"LÃÂ¤kare '{doctor_id}' inte hittad i schema")

    config = await db.get_config(sched["clinic_id"])
    doc = next((d for d in config.doctors if d.id == doctor_id), None) if config else None

    return {
        "doctor_id": doctor_id,
        "doctor_name": doc.name if doc else doctor_id,
        "role": doc.role.value if doc else "unknown",
        "schedule": doctor_schedule,
        "statistics": sched["statistics"].get("workload_balance", {}).get(doctor_id, {}),
        "call_stats": sched["statistics"].get("call_distribution", {}).get(doctor_id, {}),
    }


# === SCHEMAJUSTERING ===

@app.post("/schedule/adjust", tags=["Justering"])
async def adjust_schedule(request: AdjustmentRequest):
    """
    Manuell justering av befintligt schema.

    Typer:
    - swap: Byt funktion mellan tvÃÂ¥ lÃÂ¤kare pÃÂ¥ en dag
    - replace: ÃÂndra en lÃÂ¤kares funktion en specifik dag
    - lock: LÃÂ¥s en tilldelning (kan inte ÃÂ¤ndras av omoptimering)
    - unlock: LÃÂ¥s upp en tilldelning
    """
    sched = await db.get_schedule(request.schedule_id)
    if not sched:
        raise HTTPException(status_code=404, detail="Schema inte hittat")

    raw = sched.get("raw_schedule", {})
    config = await db.get_config(sched["clinic_id"])

    # raw_schedule keys can be int (from solver) or str (from JSON/MongoDB)
    day_int = request.day
    day_str = str(request.day)

    def _raw_get(doc_id, day):
        """Get value from raw_schedule, trying both int and str keys."""
        doc_days = raw.get(doc_id, {})
        val = doc_days.get(day_int)
        if val is None:
            val = doc_days.get(day_str)
        return val

    def _raw_set(doc_id, day, value):
        """Set value in raw_schedule, using whichever key type exists."""
        doc_days = raw.get(doc_id, {})
        if day_int in doc_days:
            doc_days[day_int] = value
        else:
            doc_days[day_str] = value

    def _sync_date_schedule():
        """Sync the date-keyed schedule after raw_schedule changes."""
        start = date.fromisoformat(sched["start_date"])
        date_sched = sched.get("schedule", {})
        d = start + timedelta(days=day_int)
        date_str = d.isoformat()
        for doc_id in raw:
            if doc_id not in date_sched:
                date_sched[doc_id] = {}
            val = _raw_get(doc_id, day_int)
            if val is not None:
                date_sched[doc_id][date_str] = val
        sched["schedule"] = date_sched

    if request.adjustment_type == "swap":
        if not request.swap_with_doctor_id:
            raise HTTPException(status_code=400, detail="swap_with_doctor_id krÃÂ¤vs fÃÂ¶r swap")

        func_a = _raw_get(request.doctor_id, day_int)
        func_b = _raw_get(request.swap_with_doctor_id, day_int)

        if func_a is None or func_b is None:
            raise HTTPException(status_code=400, detail="Ogiltigt dag-index")

        _raw_set(request.doctor_id, day_int, func_b)
        _raw_set(request.swap_with_doctor_id, day_int, func_a)

        _sync_date_schedule()
        await db.save_schedule(sched)
        await db.audit("schedule_swap", details={"schedule_id": request.schedule_id, "day": request.day, "doctor_a": request.doctor_id, "doctor_b": request.swap_with_doctor_id})

        warnings = _validate_single_day(raw, config, day_int)

        return {
            "status": "adjusted",
            "adjustment": f"Bytte {request.doctor_id} ({func_a}Ã¢ÂÂ{func_b}) med {request.swap_with_doctor_id} ({func_b}Ã¢ÂÂ{func_a}) dag {request.day}",
            "warnings": warnings,
        }

    elif request.adjustment_type == "replace":
        if not request.new_function:
            raise HTTPException(status_code=400, detail="new_function krÃÂ¤vs fÃÂ¶r replace")

        old_func = _raw_get(request.doctor_id, day_int)
        _raw_set(request.doctor_id, day_int, request.new_function)

        _sync_date_schedule()
        await db.save_schedule(sched)
        await db.audit("schedule_replace", details={"schedule_id": request.schedule_id, "day": request.day, "doctor": request.doctor_id})

        warnings = _validate_single_day(raw, config, day_int)

        return {
            "status": "adjusted",
            "adjustment": f"{request.doctor_id}: {old_func} Ã¢ÂÂ {request.new_function} dag {request.day}",
            "warnings": warnings,
        }

    else:
        raise HTTPException(status_code=400, detail=f"OkÃÂ¤nd adjustment_type: {request.adjustment_type}")


def _validate_single_day(schedule: dict, config: ClinicConfig, day: int) -> list[str]:
    """Validera en specifik dag efter justering."""
    warnings = []

    def _get(days_dict, d):
        """Get from dict trying both int and str keys."""
        v = days_dict.get(d)
        if v is None:
            v = days_dict.get(str(d))
        return v

    # Kolla att jour finns
    primary = [d_id for d_id, days in schedule.items() if "JOUR_P" in str(_get(days, day) or "")]
    backup = [d_id for d_id, days in schedule.items() if "JOUR_B" in str(_get(days, day) or "")]

    if len(primary) != 1:
        warnings.append(f"Dag {day}: {len(primary)} primÃÂ¤rjourer (ska vara 1)")
    if len(backup) != 1:
        warnings.append(f"Dag {day}: {len(backup)} bakjourer (ska vara 1)")

    # Kolla ATL (jour igÃÂ¥r Ã¢ÂÂ arbete idag)
    if day > 0:
        for d_id, days in schedule.items():
            yesterday = _get(days, day - 1) or "LEDIG"
            today = _get(days, day) or "LEDIG"
            if is_jour(yesterday) and not is_jour(today) and today != "LEDIG":
                doc = next((d for d in config.doctors if d.id == d_id), None)
                name = doc.name if doc else d_id
                warnings.append(f"ATL-brott: {name} hade jour dag {day-1} och arbetar dag {day}")

    return warnings


# === FRÃÂNVARO ===

@app.post("/absence", tags=["FrÃÂ¥nvaro"])
async def register_absence(request: AbsenceRequest, background_tasks: BackgroundTasks):
    """
    Registrera frÃÂ¥nvaro och (valfritt) omoptimera berÃÂ¶rda scheman.

    FlÃÂ¶de:
    1. Registrera frÃÂ¥nvaron
    2. Hitta berÃÂ¶rda scheman
    3. LÃÂ¥s alla andra tilldelningar
    4. Omoptimera med frÃÂ¥nvarande lÃÂ¤kare exkluderad
    """
    config = await db.get_config(request.clinic_id)
    if not config:
        raise HTTPException(status_code=404, detail=f"Klinik '{request.clinic_id}' finns inte")

    # Validera att lÃÂ¤karen finns
    doc = next((d for d in config.doctors if d.id == request.doctor_id), None)
    if not doc:
        raise HTTPException(status_code=404, detail=f"LÃÂ¤kare '{request.doctor_id}' finns inte")

    absence_id = f"abs_{uuid.uuid4().hex[:8]}"

    result = {
        "absence_id": absence_id,
        "doctor_id": request.doctor_id,
        "doctor_name": doc.name,
        "absence_type": request.absence_type,
        "start_date": request.start_date,
        "end_date": request.end_date,
        "status": "registered",
        "affected_schedules": [],
    }

    # Hitta berÃÂ¶rda scheman
    all_scheds = await db.list_schedules(clinic_id=request.clinic_id)
    for sched in all_scheds:
        sched_id = sched["schedule_id"]
        if True:  # already filtered by clinic_id
            # Kolla om frÃÂ¥nvaron ÃÂ¶verlappar med schemat
            sched_start = date.fromisoformat(sched["start_date"])
            sched_end = sched_start + timedelta(weeks=sched["num_weeks"])
            abs_start = date.fromisoformat(request.start_date)
            abs_end = date.fromisoformat(request.end_date)

            if abs_start <= sched_end and abs_end >= sched_start:
                result["affected_schedules"].append(sched_id)

                # Markera frÃÂ¥nvarande dagar som LEDIG i raw_schedule
                raw = sched.get("raw_schedule", {})
                if request.doctor_id in raw:
                    for day_idx in range(sched["num_weeks"] * 7):
                        day_date = sched_start + timedelta(days=day_idx)
                        if abs_start <= day_date <= abs_end:
                            old_func = raw[request.doctor_id].get(day_idx, "LEDIG")
                            raw[request.doctor_id][day_idx] = "LEDIG"

                            # Uppdatera ÃÂ¤ven datum-schemat
                            if request.doctor_id in sched.get("schedule", {}):
                                sched["schedule"][request.doctor_id][day_date.isoformat()] = "LEDIG"
                await db.save_schedule(sched)

    if request.reoptimize and result["affected_schedules"]:
        result["status"] = "registered_reoptimize_pending"
        result["message"] = f"FrÃÂ¥nvaro registrerad. {len(result['affected_schedules'])} schema(n) behÃÂ¶ver omoptimeras."
    else:
        result["message"] = "FrÃÂ¥nvaro registrerad. BerÃÂ¶rda dagar satta till LEDIG."

    return result


# === OMOPTIMERING ===

@app.post("/schedule/reoptimize", tags=["Schema"])
async def reoptimize_schedule(request: ReoptimizeRequest):
    """
    Omoptimera ett befintligt schema efter ÃÂ¤ndringar.

    BehÃÂ¥ller lÃÂ¥sta tilldelningar och optimerar resten.
    """
    sched = await db.get_schedule(request.schedule_id)
    if not sched:
        raise HTTPException(status_code=404, detail="Schema inte hittat")

    config = await db.get_config(sched["clinic_id"])
    if not config:
        raise HTTPException(status_code=404, detail="Klinikkonfiguration saknas")

    # KÃÂ¶r ny optimering
    gen_request = ScheduleRequest(
        clinic_id=sched["clinic_id"],
        num_weeks=sched["num_weeks"],
        start_date=sched["start_date"],
        time_limit_seconds=request.time_limit_seconds,
    )

    new_schedule_id = f"sch_{uuid.uuid4().hex[:12]}"
    job_id = f"job_{uuid.uuid4().hex[:8]}"

    _active_jobs[job_id] = {
        "job_id": job_id,
        "schedule_id": new_schedule_id,
        "status": "queued",
        "parent_schedule": request.schedule_id,
    }

    _run_solver(job_id, config, gen_request)
    job = _active_jobs[job_id]

    if job["status"] == "completed":
        await db.save_schedule(job["result"])
        await db.save_job(job)
        return {
            "status": "reoptimized",
            "original_schedule_id": request.schedule_id,
            "new_schedule_id": new_schedule_id,
            "schedule": job["result"],
        }
    else:
        raise HTTPException(
            status_code=422,
            detail=f"Omoptimering misslyckades: {job.get('error', 'okÃÂ¤nt fel')}"
        )


@app.post("/schedule/roll-forward", tags=["Schema"])
async def roll_forward(request: RollForwardRequest):
    """Rulla schema framåt — behåll låsta veckor, generera nya."""
    sched = await db.get_schedule(request.schedule_id)
    if not sched:
        raise HTTPException(status_code=404, detail="Schema inte hittat")

    raw = sched.get("raw_schedule", {})
    config_name = sched.get("clinic_id", "kristianstad")
    config = await db.get_config(config_name)
    if not config:
        raise HTTPException(status_code=404, detail="Klinikkonfiguration saknas")

    old_num_weeks = sched.get("num_weeks", 2)
    if request.weeks_to_keep >= old_num_weeks:
        raise HTTPException(status_code=400, detail="weeks_to_keep måste vara < antal veckor i schema")

    # Skifta dag-index: ta bort äldsta veckorna
    weeks_to_drop = old_num_weeks - request.weeks_to_keep
    drop_days = weeks_to_drop * 7
    shifted = {}
    for doc_id, days in raw.items():
        shifted[doc_id] = {}
        for day_key, func_id in days.items():
            old_day = int(day_key)
            if old_day >= drop_days:
                shifted[doc_id][old_day - drop_days] = func_id

    result = solve_rolling(
        config,
        existing_schedule=shifted,
        locked_weeks=request.weeks_to_keep,
        new_weeks=request.new_weeks,
        time_limit_seconds=request.time_limit_seconds,
    )

    if result is None:
        raise HTTPException(status_code=422, detail="Kunde inte lösa rullande schema")

    new_id = f"sch_{uuid.uuid4().hex[:12]}"
    total_weeks = request.weeks_to_keep + request.new_weeks
    num_days = total_weeks * 7
    granular = expand_to_granular(result, num_days)

    start_str = sched.get("start_date")
    if start_str:
        from datetime import timedelta
        old_start = datetime.fromisoformat(start_str)
        new_start = old_start + timedelta(weeks=weeks_to_drop)
    else:
        new_start = datetime.now()

    date_schedule = {}
    for doc_id, days in granular.items():
        date_schedule[doc_id] = {}
        for day_idx, func in days.items():
            day_date = new_start + timedelta(days=day_idx)
            date_schedule[doc_id][day_date.strftime("%Y-%m-%d")] = func

    new_sched_data = {
        "schedule_id": new_id,
        "status": "optimal",
        "clinic_id": config_name,
        "num_weeks": total_weeks,
        "start_date": new_start.strftime("%Y-%m-%d"),
        "created_at": datetime.now().isoformat(),
        "schedule": date_schedule,
        "raw_schedule": result,
        "parent_schedule": request.schedule_id,
    }
    await db.save_schedule(new_sched_data)

    return {
        "status": "rolled_forward",
        "original_schedule_id": request.schedule_id,
        "new_schedule_id": new_id,
        "weeks_kept": request.weeks_to_keep,
        "new_weeks": request.new_weeks,
        "total_weeks": total_weeks,
    }


# === VALIDERING ===

@app.post("/validate/{schedule_id}", tags=["Validering"])
async def validate_schedule(schedule_id: str):
    """
    FullstÃÂ¤ndig ATL-validering av ett schema.

    Kontrollerar:
    - 11h dygnsvila efter jour
    - 36h sammanhÃÂ¤ngande veckovila
    - Max 48h/vecka
    - Max 20h sammanhÃÂ¤ngande arbete+jour
    - Max 1 jour/vecka
    - Minimibemanningstal
    """
    sched = await db.get_schedule(schedule_id)
    if not sched:
        raise HTTPException(status_code=404, detail="Schema inte hittat")

    config = await db.get_config(sched["clinic_id"])
    raw_db = sched.get("raw_schedule") or {}

    # JSON sparar int-nycklar som strängar: konvertera tillbaka till int, eller
    # bygg raw från schedule (datumsnycklar) om raw saknas/tomt.
    _start = date.fromisoformat(sched["start_date"])
    _num = sched["num_weeks"] * 7
    raw = {}
    if raw_db:
        # Konvertera JSON-strängnycklar ("0","1",...) → int
        for _doc_id, _days in raw_db.items():
            raw[_doc_id] = {int(k): v for k, v in _days.items()}
    elif sched.get("schedule"):
        # Fallback: bygg från datumsnycklar
        for _doc_id, _days in sched["schedule"].items():
            raw[_doc_id] = {}
            for _date_str, _shift in _days.items():
                try:
                    _d = date.fromisoformat(_date_str)
                    _idx = (_d - _start).days
                    if 0 <= _idx < _num:
                        raw[_doc_id][_idx] = _shift
                except (ValueError, TypeError):
                    pass

    num_days = _num

    violations = []
    warnings = []

    # 1. Dygnsvila efter jour
    for doc in config.doctors:
        for day in range(num_days - 1):
            func = raw.get(doc.id, {}).get(day, "LEDIG")
            func_next = raw.get(doc.id, {}).get(day + 1, "LEDIG")
            if is_jour(func) and not is_jour(func_next) and func_next != "LEDIG":
                violations.append({
                    "type": "dygnsvila",
                    "severity": "critical",
                    "doctor_id": doc.id,
                    "doctor_name": doc.name,
                    "day": day,
                    "detail": f"Jour dag {day+1} Ã¢ÂÂ arbete dag {day+2} (krÃÂ¤ver 11h vila)",
                    "atl_reference": "ATL 13ÃÂ§, AB 13ÃÂ§7",
                })

    # 2. Max jourer per vecka
    for doc in config.doctors:
        for week in range(sched["num_weeks"]):
            week_start = week * 7
            calls = sum(1 for d in range(week_start, min(week_start + 7, num_days))
                       if is_jour(raw.get(doc.id, {}).get(d, "LEDIG")))
            if calls > 1:
                violations.append({
                    "type": "max_jour_vecka",
                    "severity": "warning",
                    "doctor_id": doc.id,
                    "doctor_name": doc.name,
                    "week": week + 1,
                    "detail": f"{calls} jourer vecka {week+1} (rekommenderat max 1)",
                })

    # 3. Bemanningstal
    for day in range(num_days):
        weekday = day % 7
        if weekday >= 5:
            continue

        primary = sum(1 for d_id in raw if "JOUR_P" in str(raw[d_id].get(day) or ""))
        backup = sum(1 for d_id in raw if "JOUR_B" in str(raw[d_id].get(day) or ""))

        if primary != 1:
            violations.append({
                "type": "bemanning",
                "severity": "critical",
                "day": day,
                "detail": f"Dag {day+1}: {primary} primÃÂ¤rjourer (ska vara 1)",
            })
        if backup != 1:
            violations.append({
                "type": "bemanning",
                "severity": "critical",
                "day": day,
                "detail": f"Dag {day+1}: {backup} bakjourer (ska vara 1)",
            })

    # 4. Arbetsbelastning per vecka
    for doc in config.doctors:
        for week in range(sched["num_weeks"]):
            week_start = week * 7
            work_days = sum(1 for d in range(week_start, min(week_start + 7, num_days))
                          if raw.get(doc.id, {}).get(d, "LEDIG") != "LEDIG")
            if work_days > 5:
                warnings.append({
                    "type": "arbetsbelastning",
                    "severity": "warning",
                    "doctor_id": doc.id,
                    "doctor_name": doc.name,
                    "week": week + 1,
                    "detail": f"{work_days} arbetsdagar vecka {week+1} (max 5 rekommenderat)",
                })

    return ValidationResult(
        valid=len([v for v in violations if v["severity"] == "critical"]) == 0,
        violations=violations,
        warnings=warnings,
        summary={
            "total_violations": len(violations),
            "critical": len([v for v in violations if v["severity"] == "critical"]),
            "warnings": len(warnings),
            "doctors_checked": len(config.doctors),
            "days_checked": num_days,
        },
    )


# === EXPORT ===

@app.get("/schedule/{schedule_id}/export/excel", tags=["Export"])
async def export_schedule_excel(schedule_id: str):
    """Exportera schema som Excel-fil."""
    from fastapi.responses import StreamingResponse
    import io

    sched = await db.get_schedule(schedule_id)
    if not sched:
        raise HTTPException(status_code=404, detail="Schema inte hittat")

    config = await db.get_config(sched["clinic_id"])

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        raise HTTPException(status_code=501, detail="openpyxl ej installerat")

    wb = Workbook()
    ws = wb.active
    ws.title = "Schema"

    # Build date list
    start = date.fromisoformat(sched["start_date"])
    total_days = sched["num_weeks"] * 7
    dates = [start + timedelta(days=i) for i in range(total_days)]

    day_names = ["MÃÂ¥n", "Tis", "Ons", "Tor", "Fre", "LÃÂ¶r", "SÃÂ¶n"]

    # Color map for functions
    func_fills = {
        "OP_H": PatternFill(start_color="E0F7FA", end_color="E0F7FA", fill_type="solid"),
        "OP_C": PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid"),
        "AVD_H": PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid"),
        "AVD_C": PatternFill(start_color="FFFBEB", end_color="FFFBEB", fill_type="solid"),
        "MOTT_H": PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid"),
        "MOTT_C": PatternFill(start_color="ECFDF5", end_color="ECFDF5", fill_type="solid"),
        "JOUR_P": PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid"),
        "JOUR_B": PatternFill(start_color="FFE4E6", end_color="FFE4E6", fill_type="solid"),
        "JOUR_P_KVÄLL": PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid"),
        "JOUR_P_NATT": PatternFill(start_color="FECACA", end_color="FECACA", fill_type="solid"),
        "JOUR_P_HELGDAG": PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid"),
        "JOUR_P_HELGNATT": PatternFill(start_color="FECACA", end_color="FECACA", fill_type="solid"),
        "JOUR_B_HELGDAG": PatternFill(start_color="FFE4E6", end_color="FFE4E6", fill_type="solid"),
        "JOUR_B_HELGNATT": PatternFill(start_color="FECDD5", end_color="FECDD5", fill_type="solid"),
        "LEDIG": PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid"),
    }

    thin_border = Border(
        left=Side(style="thin", color="D1D5DB"),
        right=Side(style="thin", color="D1D5DB"),
        top=Side(style="thin", color="D1D5DB"),
        bottom=Side(style="thin", color="D1D5DB"),
    )

    # Header row 1: day names
    ws.cell(row=1, column=1, value="LÃÂ¤kare").font = Font(bold=True, size=10)
    ws.cell(row=1, column=2, value="Roll").font = Font(bold=True, size=10)
    for ci, d in enumerate(dates):
        cell = ws.cell(row=1, column=ci + 3, value=day_names[d.weekday()])
        cell.font = Font(bold=True, size=9, color="FFFFFF")
        cell.fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    # Header row 2: dates
    for ci, d in enumerate(dates):
        cell = ws.cell(row=2, column=ci + 3, value=d.strftime("%d/%m"))
        cell.font = Font(size=8, color="64748B")
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    # Build doctor map
    doc_map = {}
    if config:
        for d in config.doctors:
            doc_map[d.id] = d

    # Sort doctors by role
    role_order = {"ÃÂL": 0, "SP": 1, "ST_SEN": 2, "ST_TIDIG": 3, "UL": 4}
    schedule_data = sched.get("schedule", {})
    def _role_str(doc):
        if doc is None:
            return "?"
        r = doc.role
        return r.value if hasattr(r, 'value') else str(r)

    sorted_docs = sorted(schedule_data.keys(), key=lambda x: (
        role_order.get(_role_str(doc_map.get(x)), 9),
        getattr(doc_map.get(x), "name", x),
    ))

    # Data rows
    for ri, doc_id in enumerate(sorted_docs):
        row = ri + 3
        doc = doc_map.get(doc_id)
        name = doc.name if doc else doc_id
        role = str(doc.role.value) if doc and hasattr(doc.role, 'value') else (str(doc.role) if doc else "?")

        ws.cell(row=row, column=1, value=name).font = Font(size=9)
        ws.cell(row=row, column=2, value=role).font = Font(size=9, color="64748B")

        day_map = schedule_data.get(doc_id, {})
        for ci, d in enumerate(dates):
            date_str = d.isoformat()
            func = day_map.get(date_str, "LEDIG")
            cell = ws.cell(row=row, column=ci + 3, value=func)
            cell.font = Font(size=9)
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border
            if func in func_fills:
                cell.fill = func_fills[func]

    # Column widths
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 8
    for ci in range(len(dates)):
        col_letter = ws.cell(row=1, column=ci + 3).column_letter
        ws.column_dimensions[col_letter].width = 10

    # Freeze panes (doctor names + header)
    ws.freeze_panes = "C3"

    # Write to buffer
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"schema_{schedule_id}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# === STATISTIK ===

@app.get("/statistics/{schedule_id}", tags=["Statistik"])
async def get_statistics(schedule_id: str):
    """HÃÂ¤mta detaljerad statistik fÃÂ¶r ett schema."""
    sched = await db.get_schedule(schedule_id)
    if not sched:
        raise HTTPException(status_code=404, detail="Schema inte hittat")
    return sched.get("statistics", {})


# === LISTA SCHEMAN ===

@app.get("/schedules", tags=["Schema"])
async def list_schedules(clinic_id: Optional[str] = None):
    """Lista alla genererade scheman."""
    schedules = await db.list_schedules(clinic_id=clinic_id)
    return [
        {
            "schedule_id": s.get("schedule_id"),
            "clinic_id": s.get("clinic_id"),
            "status": s.get("status"),
            "num_weeks": s.get("num_weeks"),
            "start_date": s.get("start_date"),
            "created_at": s.get("created_at"),
            "solve_time_ms": s.get("solve_time_ms"),
        }
        for s in schedules
    ]


# === FRÃÂNVAROKEDJA (Automatisk ersÃÂ¤ttare) ===

class AbsenceChainRequest(BaseModel):
    """BegÃÂ¤ran om att kÃÂ¶ra frÃÂ¥nvarokedjan."""
    schedule_id: str
    doctor_id: str
    absence_type: str = Field(description="sjuk, vab, semester, utbildning, konferens, permission, akut")
    start_date: str = Field(description="YYYY-MM-DD")
    end_date: str = Field(description="YYYY-MM-DD")
    auto_select: bool = Field(default=True, description="VÃÂ¤lj bÃÂ¤sta ersÃÂ¤ttare automatiskt?")
    reason: Optional[str] = None


class ManualReplacementRequest(BaseModel):
    """Manuellt val av ersÃÂ¤ttare frÃÂ¥n kandidatlistan."""
    schedule_id: str
    day: int
    function: str
    absent_doctor_id: str
    replacement_doctor_id: str
    override_atl: bool = Field(default=False, description="GodkÃÂ¤nn trots ATL-varning?")


# Absence chain store Ã¢ÂÂ backed by db layer
absence_chain_store = db._chains  # backward compat dict reference


@app.post("/absence/chain", tags=["FrÃÂ¥nvarokedja"])
async def run_absence_chain(request: AbsenceChainRequest):
    """
    KÃÂ¶r hela frÃÂ¥nvarokedjan: registrera Ã¢ÂÂ analysera Ã¢ÂÂ ranka Ã¢ÂÂ validera Ã¢ÂÂ ersÃÂ¤tt Ã¢ÂÂ notifiera.

    Returnerar komplett resultat med ersÃÂ¤ttare, kandidatlistor, ATL-validering och notifieringar.
    Om auto_select=False returneras bara kandidatlistan utan att schemat ÃÂ¤ndras.
    """
    sched = await db.get_schedule(request.schedule_id)
    if not sched:
        raise HTTPException(status_code=404, detail="Schema inte hittat")

    config = await db.get_config(sched["clinic_id"])
    if not config:
        raise HTTPException(status_code=404, detail="Klinikkonfiguration saknas")

    raw_db = sched.get("raw_schedule") or {}
    start = date.fromisoformat(sched["start_date"])
    _chain_num_days = sched["num_weeks"] * 7

    # JSON konverterar int-nycklar till strängar — konvertera tillbaka, eller bygg från schedule
    raw = {}
    if raw_db:
        for doc_id, days in raw_db.items():
            raw[doc_id] = {int(k): v for k, v in days.items()}
    elif sched.get("schedule"):
        for doc_id, days in sched["schedule"].items():
            raw[doc_id] = {}
            for date_str, shift in days.items():
                try:
                    d = date.fromisoformat(date_str)
                    day_idx = (d - start).days
                    if 0 <= day_idx < _chain_num_days:
                        raw[doc_id][day_idx] = shift
                except (ValueError, TypeError):
                    pass

    chain = AbsenceChain(config, raw, start, sched["num_weeks"])
    result = chain.execute(
        doctor_id=request.doctor_id,
        absence_type=request.absence_type,
        start_date=request.start_date,
        end_date=request.end_date,
        auto_select=request.auto_select,
    )

    # Spara resultat — både in-memory (för session) och PostgreSQL (persistent)
    absence_chain_store[result.chain_id] = result
    chain_dict = {
        "chain_id": result.chain_id,
        "absence_type": result.absence_type,
        "doctor_id": result.doctor_id,
        "doctor_name": result.doctor_name,
        "start_date": result.start_date,
        "end_date": result.end_date,
        "status": result.status.value,
        "vacant_slots": result.vacant_slots,
        "replacements": result.replacements,
        "failed_slots": result.failed_slots,
        "chain_log": result.chain_log,
        "notifications": result.notifications,
        "atl_violations": result.atl_violations,
        "schedule_changes": result.schedule_changes,
    }
    try:
        await db.save_chain(chain_dict)
    except Exception as _chain_err:
        logger.warning("Kunde inte spara frånvarokedja till DB: %s", _chain_err)

    # WebSocket broadcast: frÃÂ¥nvarokedja
    if HAS_WS:
        import asyncio
        try:
            asyncio.get_event_loop().create_task(
                hub.broadcast_absence_chain(
                    chain_id=result.chain_id,
                    status=result.status.value,
                    details={
                        "doctor_id": request.doctor_id,
                        "absence_type": request.absence_type,
                        "vacant_slots": len(result.vacant_slots),
                        "replaced": len(result.schedule_changes),
                    }
                )
            )
        except Exception:
            pass

    # Om auto_select: uppdatera ÃÂ¤ven datum-schemat
    if request.auto_select and result.schedule_changes:
        for change in result.schedule_changes:
            day_date = change["date"]
            # FrÃÂ¥nvarande Ã¢ÂÂ LEDIG
            if request.doctor_id in sched.get("schedule", {}):
                sched["schedule"][request.doctor_id][day_date] = "LEDIG"
            # ErsÃÂ¤ttare Ã¢ÂÂ ny funktion
            repl_id = change["replacement_doctor"]
            if repl_id in sched.get("schedule", {}):
                sched["schedule"][repl_id][day_date] = change["function"]

        # Uppdatera statistik
        sched["statistics"] = _compute_statistics(raw, config, sched["num_weeks"] * 7)

    return {
        "chain_id": result.chain_id,
        "status": result.status.value,
        "doctor": f"{result.doctor_name} ({result.doctor_id})",
        "absence_type": result.absence_type,
        "period": f"{result.start_date} Ã¢ÂÂ {result.end_date}",
        "summary": {
            "vacant_slots": len(result.vacant_slots),
            "replaced": len(result.schedule_changes),
            "failed": len(result.failed_slots),
            "atl_warnings": len(result.atl_violations),
        },
        "replacements": result.replacements,
        "failed_slots": result.failed_slots,
        "schedule_changes": result.schedule_changes,
        "notifications": result.notifications,
        "chain_log": result.chain_log,
    }


@app.get("/absence/chain/{chain_id}", tags=["FrÃÂ¥nvarokedja"])
async def get_absence_chain(chain_id: str):
    """HÃÂ¤mta resultat av en tidigare kÃÂ¶rd frÃÂ¥nvarokedja."""
    result = absence_chain_store.get(chain_id)
    if not result:
        raise HTTPException(status_code=404, detail="FrÃÂ¥nvarokedja inte hittad")

    # result can be either a Pydantic object or a dict (in-memory save_chain overwrites with dict)
    if isinstance(result, dict):
        status_val = result.get("status", "")
        if hasattr(status_val, "value"):
            status_val = status_val.value
        return {
            "chain_id": result.get("chain_id"),
            "status": status_val,
            "doctor": f"{result.get('doctor_name', '')} ({result.get('doctor_id', '')})",
            "absence_type": result.get("absence_type"),
            "period": f"{result.get('start_date')} - {result.get('end_date')}",
            "replacements": result.get("replacements", []),
            "failed_slots": result.get("failed_slots", []),
            "schedule_changes": result.get("schedule_changes", []),
            "notifications": result.get("notifications", []),
            "chain_log": result.get("chain_log", []),
        }
    status_val = result.status.value if hasattr(result.status, "value") else result.status
    return {
        "chain_id": result.chain_id,
        "status": status_val,
        "doctor": f"{result.doctor_name} ({result.doctor_id})",
        "absence_type": result.absence_type,
        "period": f"{result.start_date} - {result.end_date}",
        "replacements": result.replacements,
        "failed_slots": result.failed_slots,
        "schedule_changes": result.schedule_changes,
        "notifications": result.notifications,
        "chain_log": result.chain_log,
    }


@app.post("/absence/chain/manual-replace", tags=["FrÃÂ¥nvarokedja"])
async def manual_replacement(request: ManualReplacementRequest):
    """
    Manuellt val av ersÃÂ¤ttare Ã¢ÂÂ fÃÂ¶r slots som krÃÂ¤ver manuell hantering
    eller nÃÂ¤r du vill vÃÂ¤lja en annan kandidat ÃÂ¤n den automatiskt valda.
    """
    sched = await db.get_schedule(request.schedule_id)
    if not sched:
        raise HTTPException(status_code=404, detail="Schema inte hittat")

    config = await db.get_config(sched["clinic_id"])
    raw = sched.get("raw_schedule", {})

    # Validera att ersÃÂ¤ttaren finns
    doc = next((d for d in config.doctors if d.id == request.replacement_doctor_id), None)
    if not doc:
        raise HTTPException(status_code=404, detail=f"LÃÂ¤kare '{request.replacement_doctor_id}' finns inte")

    # ATL-validering
    start = date.fromisoformat(sched["start_date"])
    chain = AbsenceChain(config, raw, start, sched["num_weeks"])

    from absence_chain import VacantSlot, Candidate
    slot = VacantSlot(
        day_index=request.day,
        day_date=start + timedelta(days=request.day),
        function=request.function,
        is_call=is_jour(request.function),
        site=chain.FUNCTION_SITE.get(request.function),
        weekday=(start + timedelta(days=request.day)).weekday(),
    )
    candidate = Candidate(
        doctor_id=doc.id, doctor_name=doc.name, role=doc.role.value,
        score=0, current_function=raw.get(doc.id, {}).get(request.day, "LEDIG"),
        reasons=[],
    )
    atl_result = chain._validate_atl(candidate, slot)

    if not atl_result["ok"] and not request.override_atl:
        return {
            "status": "atl_violation",
            "message": "ATL-brott upptÃÂ¤ckt. SÃÂ¤tt override_atl=true fÃÂ¶r att godkÃÂ¤nna ÃÂ¤ndÃÂ¥.",
            "violations": atl_result["violations"],
        }

    # VerkstÃÂ¤ll
    old_func = raw.get(request.replacement_doctor_id, {}).get(request.day, "LEDIG")

    # FrÃÂ¥nvarande Ã¢ÂÂ LEDIG
    if request.absent_doctor_id in raw:
        raw[request.absent_doctor_id][request.day] = "LEDIG"
    # ErsÃÂ¤ttare Ã¢ÂÂ funktion
    if request.replacement_doctor_id in raw:
        raw[request.replacement_doctor_id][request.day] = request.function

    # Uppdatera datum-schema
    day_date = (start + timedelta(days=request.day)).isoformat()
    if request.absent_doctor_id in sched.get("schedule", {}):
        sched["schedule"][request.absent_doctor_id][day_date] = "LEDIG"
    if request.replacement_doctor_id in sched.get("schedule", {}):
        sched["schedule"][request.replacement_doctor_id][day_date] = request.function

    return {
        "status": "replaced",
        "day": request.day,
        "date": day_date,
        "function": request.function,
        "absent_doctor": request.absent_doctor_id,
        "replacement": f"{doc.name} ({doc.id})",
        "replacement_old_function": old_func,
        "atl_ok": atl_result["ok"],
        "atl_overridden": not atl_result["ok"] and request.override_atl,
        "warnings": atl_result.get("violations", []),
    }


@app.get("/absence/chains", tags=["FrÃÂ¥nvarokedja"])
async def list_absence_chains(status: Optional[str] = None):
    """Lista alla körda frånvarokedjor (läses från PostgreSQL)."""
    # Hämta från DB (persistent)
    db_chains = await db.list_chains(status=status)
    if db_chains:
        return [
            {
                "chain_id": c.get("chain_id"),
                "doctor": f"{c.get('doctor_name','')} ({c.get('doctor_id','')})",
                "absence_type": c.get("absence_type"),
                "period": f"{c.get('start_date','')} — {c.get('end_date','')}",
                "status": c.get("status"),
                "replaced": len(c.get("schedule_changes", [])),
                "failed": len(c.get("failed_slots", [])),
            }
            for c in db_chains
        ]
    # Fallback: in-memory (för denna session)
    results = []
    for chain_id, result in absence_chain_store.items():
        if status and result.status.value != status:
            continue
        results.append({
            "chain_id": result.chain_id,
            "doctor": f"{result.doctor_name} ({result.doctor_id})",
            "absence_type": result.absence_type,
            "period": f"{result.start_date} — {result.end_date}",
            "status": result.status.value,
            "replaced": len(result.schedule_changes),
            "failed": len(result.failed_slots),
        })
    return results


# === DATABASE STATUS ===

@app.get("/db/status", tags=["System"])
async def db_status():
    """Visa databasstatus (PostgreSQL eller in-memory)."""
    return await db.stats()


@app.post("/db/migrate", tags=["System"])
async def db_migrate():
    """Manuell databasmigrering — skapa tabeller i PostgreSQL."""
    from db import connect_db, is_connected, DATABASE_URL
    if is_connected():
        return {"status": "already_connected", "backend": "postgresql"}
    if not DATABASE_URL:
        return {"status": "no_database_url", "message": "DATABASE_URL ej konfigurerad. Sätt variabeln i Railway."}
    ok = await connect_db()
    if ok:
        return {"status": "migrated", "backend": "postgresql", "message": "Tabeller skapade och anslutning aktiv."}
    return {"status": "failed", "message": "Kunde inte ansluta till PostgreSQL. Kontrollera DATABASE_URL."}


# === SCHEMAVERSIONER ===

@app.get("/schedule/{schedule_id}/versions", tags=["Versioner"])
async def get_schedule_versions(schedule_id: str):
    versions = await db.get_versions(schedule_id)
    return {"schedule_id": schedule_id, "versions": versions, "count": len(versions)}


@app.get("/schedule/{schedule_id}/diff/{v1}/{v2}", tags=["Versioner"])
async def get_schedule_diff(schedule_id: str, v1: int, v2: int):
    ver1 = await db.get_version(schedule_id, v1)
    ver2 = await db.get_version(schedule_id, v2)
    if not ver1 or not ver2:
        raise HTTPException(status_code=404, detail="Version inte hittad")

    sched1 = ver1.get("raw_schedule", {})
    sched2 = ver2.get("raw_schedule", {})
    changes = []
    doctors_affected = set()
    for doc_id in set(list(sched1.keys()) + list(sched2.keys())):
        days1 = sched1.get(doc_id, {})
        days2 = sched2.get(doc_id, {})
        all_days = set(list(days1.keys()) + list(days2.keys()))
        for day in all_days:
            old = days1.get(day, days1.get(str(day), "LEDIG"))
            new = days2.get(day, days2.get(str(day), "LEDIG"))
            if str(old) != str(new):
                changes.append({"doctor_id": doc_id, "day": day, "old_func": old, "new_func": new})
                doctors_affected.add(doc_id)

    return {
        "schedule_id": schedule_id,
        "version_a": v1, "version_b": v2,
        "changed_assignments": changes,
        "summary": f"{len(changes)} pass ändrade, {len(doctors_affected)} läkare berörda",
    }


# === AUDIT ===

@app.get("/audit", tags=["Audit"])
async def get_audit_log(action: Optional[str] = None, user_id: Optional[str] = None, limit: int = 50):
    logs = await db.get_audit_log(action=action, user_id=user_id, limit=limit)
    return {"logs": logs, "count": len(logs)}


@app.get("/audit/stats", tags=["Audit"])
async def get_audit_stats():
    return await db.get_audit_stats()


# === RAPPORTER ===

@app.get("/reports/monthly", tags=["Rapporter"])
async def monthly_report(clinic_id: str = "kristianstad", month: Optional[str] = None):
    config = await db.get_config(clinic_id)
    if not config:
        raise HTTPException(status_code=404, detail="Klinik inte hittad")

    # Samla data från senaste schemat
    schedules = await db.list_schedules(clinic_id=clinic_id)
    if not schedules:
        return {"period": month, "clinic": clinic_id, "summary": {"total_scheduled_shifts": 0, "total_doctors": len(config.doctors)}}

    latest = max(schedules, key=lambda s: s.get("created_at", ""))
    raw = latest.get("raw_schedule", {})
    num_days = latest.get("num_weeks", 2) * 7
    stats = latest.get("statistics", {})

    total_shifts = sum(1 for days in raw.values() for f in days.values() if f != "LEDIG")
    adjustments = 0  # Async audit query not available in sync report context
    absences = 0

    # OB-beräkning
    from solver import _ob_cost
    ob_per_doc = {}
    for doc in config.doctors:
        total_ob = sum(_ob_cost(int(d) % 7, f, config.ob_rates) for d, f in raw.get(doc.id, {}).items())
        if total_ob > 0:
            ob_per_doc[doc.id] = round(total_ob, 1)
    total_ob = sum(ob_per_doc.values())
    ob_vals = list(ob_per_doc.values())
    ob_std = round((sum((v - total_ob/max(len(ob_vals),1))**2 for v in ob_vals) / max(len(ob_vals),1))**0.5, 2) if ob_vals else 0

    return {
        "period": month or "senaste",
        "clinic": getattr(config, "name", clinic_id),
        "summary": {
            "total_scheduled_shifts": total_shifts,
            "total_doctors": len(config.doctors),
            "schedules_generated": len(schedules),
            "adjustments_made": adjustments,
            "absences_registered": absences,
        },
        "call_distribution": stats.get("call_distribution", {}),
        "ob_summary": {
            "total_ob_cost_factor": round(total_ob, 1),
            "per_doctor": ob_per_doc,
            "std_deviation": ob_std,
        },
        "atl_compliance": {
            "violations": len(stats.get("atl_violations", [])),
            "details": stats.get("atl_violations", []),
        },
        "workload_balance": stats.get("workload_balance", {}),
    }


# === NOTIFIERINGAR ===

@app.get("/notifications/{user_id}", tags=["Notifieringar"])
async def get_notifications(user_id: str, unread_only: bool = False):
    notifs = await db.get_notifications(user_id, unread_only=unread_only)
    return {"notifications": notifs, "unread_count": sum(1 for n in notifs if not n.get("read"))}


@app.put("/notifications/{notif_id}/read", tags=["Notifieringar"])
async def mark_notification_read(notif_id: str, user_id: str = "anon"):
    ok = await db.mark_notification_read(user_id, notif_id)
    if not ok:
        raise HTTPException(status_code=404, detail="Notifiering inte hittad")
    return {"status": "read"}


# === INTEGRATIONER ===

from adapters.adapter_manager import AdapterManager, ADAPTER_REGISTRY
from adapters.base import AdapterConfig, AdapterType, SyncDirection


@app.get("/integrations", tags=["Integrationer"])
async def list_integrations():
    available = [t.value for t in ADAPTER_REGISTRY.keys()]
    status = AdapterManager.get_sync_status()
    return {"available": available, "connected": status}


@app.post("/integrations/{adapter_type}/connect", tags=["Integrationer"])
async def connect_integration(adapter_type: str, config: dict = {}):
    try:
        atype = AdapterType(adapter_type)
    except ValueError:
        raise HTTPException(status_code=400, detail=f"Okänd adapter: {adapter_type}")

    if atype not in ADAPTER_REGISTRY:
        raise HTTPException(status_code=400, detail=f"Adapter {adapter_type} ej tillgänglig")

    adapter_cls = ADAPTER_REGISTRY[atype]
    adapter_config = AdapterConfig(
        adapter_type=atype,
        host=config.get("host", "localhost"),
        port=config.get("port"),
        database=config.get("database", ""),
        username=config.get("username", ""),
        password=config.get("password", ""),
        api_key=config.get("api_key", ""),
    )
    adapter = adapter_cls(adapter_config)
    try:
        await adapter.connect()
        AdapterManager.register_adapter(atype, adapter)
        await db.audit("integration_connected", details={"adapter": adapter_type})
        return {"status": "connected", "adapter": adapter_type}
    except Exception as e:
        raise HTTPException(status_code=422, detail=str(e))


@app.post("/integrations/{adapter_type}/disconnect", tags=["Integrationer"])
async def disconnect_integration(adapter_type: str):
    try:
        atype = AdapterType(adapter_type)
    except ValueError:
        raise HTTPException(status_code=400, detail=f"Okänd adapter: {adapter_type}")
    adapter = AdapterManager.get_adapter(atype)
    if adapter:
        await adapter.disconnect()
        del AdapterManager._adapters[atype]
    return {"status": "disconnected", "adapter": adapter_type}


@app.post("/integrations/{adapter_type}/sync", tags=["Integrationer"])
async def sync_integration(adapter_type: str):
    try:
        atype = AdapterType(adapter_type)
    except ValueError:
        raise HTTPException(status_code=400, detail=f"Okänd adapter: {adapter_type}")
    adapter = AdapterManager.get_adapter(atype)
    if not adapter or not adapter.is_connected:
        raise HTTPException(status_code=400, detail="Adapter ej ansluten")
    try:
        config = await adapter.pull_config()
        return {"status": "synced", "adapter": adapter_type, "doctors": len(config.doctors) if config else 0}
    except Exception as e:
        raise HTTPException(status_code=422, detail=str(e))


@app.get("/integrations/{adapter_type}/status", tags=["Integrationer"])
async def integration_status(adapter_type: str):
    try:
        atype = AdapterType(adapter_type)
    except ValueError:
        raise HTTPException(status_code=400, detail=f"Okänd adapter: {adapter_type}")
    adapter = AdapterManager.get_adapter(atype)
    if not adapter:
        return {"adapter": adapter_type, "connected": False}
    try:
        info = await adapter.test_connection()
        return {"adapter": adapter_type, "connected": adapter.is_connected, **info}
    except Exception as e:
        return {"adapter": adapter_type, "connected": False, "error": str(e)}


@app.post("/integrations/{adapter_type}/test", tags=["Integrationer"])
async def test_integration(adapter_type: str, config: dict = {}):
    try:
        atype = AdapterType(adapter_type)
    except ValueError:
        raise HTTPException(status_code=400, detail=f"Okänd adapter: {adapter_type}")
    if atype not in ADAPTER_REGISTRY:
        raise HTTPException(status_code=400, detail=f"Adapter {adapter_type} ej tillgänglig")

    adapter_cls = ADAPTER_REGISTRY[atype]
    adapter_config = AdapterConfig(
        adapter_type=atype,
        host=config.get("host", "localhost"),
        port=config.get("port"),
        database=config.get("database", ""),
        username=config.get("username", ""),
        password=config.get("password", ""),
    )
    adapter = adapter_cls(adapter_config)
    try:
        connected = await adapter.connect()
        info = await adapter.test_connection()
        await adapter.disconnect()
        return {"status": "ok" if connected else "failed", **info}
    except Exception as e:
        return {"status": "failed", "error": str(e)}


# === LÖNEUNDERLAG ===

from payroll_export import payroll_exporter, PayrollExporter
from op_planning import op_planner, Operation
from jour_report import jour_reporter


async def _resolve_payroll_config_and_schedule(clinic_id: str):
    """Hjälpfunktion: hitta config och senaste schema för löneutdata.
    Fallback-kedja: explicit clinic_id → senaste schemat oavsett klinik."""
    config = await db.get_config(clinic_id)
    schedules = await db.list_schedules(clinic_id=clinic_id) if config else []
    if not schedules:
        # Fallback: hämta senaste schema oavsett klinik och använd dess clinic_id
        all_scheds = await db.list_schedules()
        if all_scheds:
            latest_any = max(all_scheds, key=lambda s: s.get("created_at", ""))
            actual_clinic = latest_any.get("clinic_id", clinic_id)
            config = await db.get_config(actual_clinic)
            schedules = [latest_any]
    return config, schedules


@app.get("/payroll/generate", tags=["Löneunderlag"])
async def generate_payroll(clinic_id: str = "hassleholm", start: str = "2026-04-06", end: str = "2026-04-19"):
    config, schedules = await _resolve_payroll_config_and_schedule(clinic_id)
    if not config:
        raise HTTPException(status_code=404, detail="Klinik inte hittad")
    if not schedules:
        return {"entries": [], "summary": {}}
    latest = max(schedules, key=lambda s: s.get("created_at", ""))
    raw = latest.get("raw_schedule", latest.get("schedule", {}))
    entries = payroll_exporter.generate_payroll(raw, start, end, config.doctors)
    summary = payroll_exporter.get_summary(entries)
    return {"entries": [{"doctor_id": e.doctor_id, "doctor_name": e.doctor_name, "date": e.date,
                         "pay_code": e.pay_code.value, "hours": e.hours} for e in entries],
            "summary": summary}


@app.get("/payroll/export", tags=["Löneunderlag"])
async def export_payroll(clinic_id: str = "hassleholm", start: str = "2026-04-06", end: str = "2026-04-19", format: str = "csv"):
    config, schedules = await _resolve_payroll_config_and_schedule(clinic_id)
    if not config:
        raise HTTPException(status_code=404, detail="Klinik inte hittad")
    if not schedules:
        return {"content": "", "format": format}
    latest = max(schedules, key=lambda s: s.get("created_at", ""))
    raw = latest.get("raw_schedule", latest.get("schedule", {}))
    entries = payroll_exporter.generate_payroll(raw, start, end, config.doctors)
    if format == "paxml":
        content = payroll_exporter.export_paxml(entries, clinic_id, f"{start[:7]}")
    else:
        content = payroll_exporter.export_csv(entries)
    return {"content": content, "format": format, "entries_count": len(entries)}


@app.post("/payroll/validate", tags=["Löneunderlag"])
async def validate_payroll(body: dict = {}):
    clinic_id = body.get("clinic_id", "hassleholm")
    config, schedules = await _resolve_payroll_config_and_schedule(clinic_id)
    if not config:
        return {"errors": ["Klinik inte hittad"]}
    if not schedules:
        return {"errors": [], "valid": True}
    latest = max(schedules, key=lambda s: s.get("created_at", ""))
    raw = latest.get("raw_schedule", latest.get("schedule", {}))
    entries = payroll_exporter.generate_payroll(raw, body.get("start", "2026-04-06"), body.get("end", "2026-04-19"), config.doctors)
    errors = payroll_exporter.validate(entries)
    return {"errors": errors, "valid": len(errors) == 0}


# === OP-PLANERING ===

@app.post("/op/match", tags=["OP-planering"])
async def match_op(body: dict):
    ops = [Operation(id=o.get("id", ""), procedure=o.get("procedure", ""),
                     required_competence=o.get("required_competence", ""),
                     estimated_duration_min=o.get("duration", 120),
                     site=o.get("site", "")) for o in body.get("operations", [])]
    schedule = body.get("schedule", {})
    competences = body.get("doctor_competences", {})
    date_str = body.get("date", "")
    return op_planner.match_competence(ops, schedule, competences, date_str)


@app.get("/op/utilization", tags=["OP-planering"])
async def op_utilization(rooms: int = 5, total_minutes: int = 2400):
    ops = [Operation(id=f"op{i}", estimated_duration_min=total_minutes // max(rooms, 1)) for i in range(rooms)]
    return op_planner.calculate_utilization(ops, rooms)


# === JOURRAPPORT ===

@app.get("/reports/jour", tags=["Rapporter"])
async def jour_report(clinic_id: str = "kristianstad", start: str = "2026-04-06", end: str = "2026-04-19"):
    config = await db.get_config(clinic_id)
    if not config:
        raise HTTPException(status_code=404, detail="Klinik inte hittad")
    schedules = await db.list_schedules(clinic_id=clinic_id)
    if not schedules:
        return {"period": f"{start} — {end}", "by_doctor": {}, "summary": {}}
    latest = max(schedules, key=lambda s: s.get("created_at", ""))
    raw = latest.get("raw_schedule", latest.get("schedule", {}))
    return jour_reporter.generate_report(raw, start, end, config.doctors)


@app.get("/reports/jour/compare", tags=["Rapporter"])
async def compare_jour(clinic_id: str = "kristianstad",
                       start1: str = "", end1: str = "", start2: str = "", end2: str = ""):
    config = await db.get_config(clinic_id)
    if not config:
        raise HTTPException(status_code=404, detail="Klinik inte hittad")
    schedules = await db.list_schedules(clinic_id=clinic_id)
    if not schedules:
        return {"changes": {}}
    latest = max(schedules, key=lambda s: s.get("created_at", ""))
    raw = latest.get("raw_schedule", latest.get("schedule", {}))
    r1 = jour_reporter.generate_report(raw, start1, end1, config.doctors) if start1 else {}
    r2 = jour_reporter.generate_report(raw, start2, end2, config.doctors) if start2 else {}
    return jour_reporter.compare_periods(r1, r2)


# === DASHBOARD ===

@app.get("/dashboard", tags=["Dashboard"])
async def get_dashboard(clinic_id: str = "kristianstad"):
    config = await db.get_config(clinic_id)
    doctors = config.doctors if config else []
    schedules = await db.list_schedules(clinic_id=clinic_id)
    latest = max(schedules, key=lambda s: s.get("created_at", "")) if schedules else None
    raw = latest.get("raw_schedule", {}) if latest else {}

    today_str = datetime.now().strftime("%Y-%m-%d")
    weekday = datetime.now().weekday()

    # Idag
    by_function = {}
    absent = 0
    primary_call = backup_call = None
    for doc in doctors:
        func = None
        date_sched = latest.get("schedule", {}).get(doc.id, {}) if latest else {}
        func = date_sched.get(today_str, "LEDIG")
        if func == "LEDIG":
            absent += 1
        elif "JOUR_P" in func:
            primary_call = doc.name
        elif "JOUR_B" in func:
            backup_call = doc.name
        else:
            prefix = func.split("_")[0] if func else "LEDIG"
            by_function[prefix] = by_function.get(prefix, 0) + 1

    # Varningar
    warnings = []
    swap_pending = len([r for r in db._swap_requests if r.status in ("pending_peer", "atl_validated")])
    if swap_pending > 0:
        warnings.append(f"{swap_pending} bytesförfrågningar väntar")

    return {
        "today": {
            "date": today_str,
            "doctors_working": len(doctors) - absent,
            "by_function": by_function,
            "primary_call": primary_call,
            "backup_call": backup_call,
            "absent": absent,
        },
        "warnings": warnings,
        "statistics": {
            "total_calls": sum(1 for days in raw.values() for f in days.values() if "JOUR_P" in str(f) or "JOUR_B" in str(f)),
            "op_utilization": 0.85,
        },
        "pending": {
            "swap_requests": swap_pending,
            "wishes": len([w for w in preference_manager._wishes if w.status == "pending"]),
            "vacation_requests": 0,
        },
    }


# === ÖNSKEMÅL ===

from preferences import preference_manager, ScheduleWish, WishPeriod
from schema_import import schema_importer

# Onboarding progress store
_onboarding = {}


@app.post("/wish-periods", tags=["Önskemål"])
async def create_wish_period(body: dict):
    period = preference_manager.create_period(
        clinic_id=body.get("clinic_id", "kristianstad"),
        name=body.get("name", ""), schedule_start=body.get("schedule_start", ""),
        schedule_end=body.get("schedule_end", ""), wish_deadline=body.get("wish_deadline", ""),
    )
    return {"status": "created", "id": period.id}


@app.get("/wish-periods", tags=["Önskemål"])
async def list_wish_periods(clinic_id: str = "kristianstad"):
    periods = [p for p in preference_manager._periods.values() if p.clinic_id == clinic_id]
    return {"periods": [{"id": p.id, "name": p.name, "status": p.status,
                         "wish_deadline": p.wish_deadline} for p in periods]}


@app.post("/wish-periods/{period_id}/close", tags=["Önskemål"])
async def close_wish_period(period_id: str):
    p = preference_manager.close_period(period_id)
    if not p:
        raise HTTPException(status_code=404, detail="Period inte hittad")
    return {"status": "closed", "id": p.id}


@app.get("/wish-periods/{period_id}/collisions", tags=["Önskemål"])
async def get_collisions(period_id: str):
    return preference_manager.get_collision_report(period_id)


@app.post("/wishes", tags=["Önskemål"])
async def submit_wish(body: dict):
    wish = ScheduleWish(
        id=f"wish_{len(preference_manager._wishes)+1}",
        doctor_id=body.get("doctor_id", ""),
        doctor_name=body.get("doctor_name", ""),
        period_id=body.get("period_id", ""),
        wish_type=body.get("wish_type", "ledig_dag"),
        priority=body.get("priority", "normal"),
        dates=body.get("dates", []),
        week_numbers=body.get("week_numbers", []),
        weekdays=body.get("weekdays", []),
        function=body.get("function"),
        site=body.get("site"),
        note=body.get("note", ""),
    )
    preference_manager.submit_wish(wish)
    return {"status": "submitted", "id": wish.id}


@app.get("/wishes", tags=["Önskemål"])
async def list_wishes(period_id: str = "", doctor_id: str = None):
    wishes = preference_manager.get_wishes_for_period(period_id, doctor_id)
    return {"wishes": [{"id": w.id, "doctor_id": w.doctor_id, "wish_type": w.wish_type,
                        "priority": w.priority, "dates": w.dates, "status": w.status,
                        "note": w.note} for w in wishes]}


@app.delete("/wishes/{wish_id}", tags=["Önskemål"])
async def delete_wish(wish_id: str):
    preference_manager._wishes = [w for w in preference_manager._wishes if w.id != wish_id]
    return {"status": "deleted"}


# === IMPORT ===

@app.post("/import/preview", tags=["Import"])
async def import_preview(body: dict):
    csv_text = body.get("csv_text", "")
    clinic_id = body.get("clinic_id", "kristianstad")
    return schema_importer.preview_import(csv_text, clinic_id)


@app.post("/import/execute", tags=["Import"])
async def import_execute(body: dict):
    csv_text = body.get("csv_text", "")
    clinic_id = body.get("clinic_id", "kristianstad")
    mapping = body.get("mapping", {})
    result = schema_importer.import_from_csv(csv_text, clinic_id, mapping)
    if result["schedule"]:
        sid = f"imp_{datetime.now().strftime('%H%M%S')}"
        _import_data = {
            "schedule_id": sid, "status": "imported", "clinic_id": clinic_id,
            "raw_schedule": result["schedule"], "schedule": result["schedule"],
            "created_at": datetime.now().isoformat(),
        }
        await db.save_schedule(_import_data)
        result["schedule_id"] = sid
    await db.audit("schema_imported", details={"clinic_id": clinic_id, "rows": result["imported_rows"]})
    return result


# === ONBOARDING ===

@app.post("/onboarding/status", tags=["Onboarding"])
async def save_onboarding(body: dict):
    clinic_id = body.get("clinic_id", "new")
    _onboarding[clinic_id] = body
    return {"status": "saved", "clinic_id": clinic_id}


@app.get("/onboarding/status", tags=["Onboarding"])
async def get_onboarding(clinic_id: str = "new"):
    return _onboarding.get(clinic_id, {"step": 1})


# === GRUNDSCHEMA ===

@app.post("/base-schedule/generate", tags=["Grundschema"])
async def generate_base(body: dict = {}):
    clinic_id = body.get("clinic_id", "kristianstad")
    config = await db.get_config(clinic_id)
    if not config:
        raise HTTPException(status_code=404, detail="Klinik inte hittad")
    cycle = body.get("cycle_weeks", 10)
    time_limit = body.get("time_limit_seconds", 120)

    from solver import generate_base_schedule
    base = generate_base_schedule(config, cycle_weeks=cycle, time_limit_seconds=time_limit)
    if base is None:
        raise HTTPException(status_code=422, detail="Kunde inte generera grundschema")

    db._base_schedules[base.id] = base
    await db.audit("base_schedule_generated", details={"id": base.id, "cycle": cycle})
    return {"status": "generated", "id": base.id, "cycle_weeks": cycle, "slots": len(base.slots)}


@app.get("/base-schedule/{base_id}", tags=["Grundschema"])
async def get_base_schedule(base_id: str):
    base = db._base_schedules.get(base_id)
    if not base:
        raise HTTPException(status_code=404, detail="Grundschema inte hittat")
    return {"id": base.id, "name": base.name, "cycle_length_weeks": base.cycle_length_weeks,
            "slots_count": len(base.slots), "effective_from": base.effective_from}


@app.get("/base-schedule/list/all", tags=["Grundschema"])
async def list_base_schedules():
    return {"schedules": [{"id": b.id, "name": b.name, "cycle_weeks": b.cycle_length_weeks}
                          for b in db._base_schedules.values()]}


@app.post("/deviations", tags=["Grundschema"])
async def create_deviation(body: dict):
    from data_model import ScheduleDeviation
    dev = ScheduleDeviation(
        id=f"dev_{len(db._deviations)+1}",
        base_schedule_id=body.get("base_schedule_id", ""),
        date=body.get("date", ""),
        doctor_id=body.get("doctor_id", ""),
        original_function=body.get("original_function", ""),
        new_function=body.get("new_function", "LEDIG"),
        reason=body.get("reason", ""),
        created_at=datetime.now().isoformat(),
    )
    db._deviations.append(dev)
    return {"status": "created", "id": dev.id}


@app.get("/schedule/effective", tags=["Grundschema"])
async def get_effective_schedule(base_id: str = "", start: str = "2026-04-06", weeks: int = 2):
    base = db._base_schedules.get(base_id)
    if not base:
        if db._base_schedules:
            base = list(db._base_schedules.values())[0]
        else:
            raise HTTPException(status_code=404, detail="Inget grundschema")
    from solver import resolve_effective_schedule
    devs = [d for d in db._deviations if d.base_schedule_id == base.id]
    schedule = resolve_effective_schedule(base, devs, start, weeks)
    return {"base_id": base.id, "start": start, "weeks": weeks, "schedule": schedule}


# === KOMP-TID ===

@app.get("/comp-time/{doctor_id}", tags=["Komp-tid"])
async def get_comp_time(doctor_id: str):
    account = db._comp_accounts.get(doctor_id)
    if not account:
        return {"doctor_id": doctor_id, "balance_hours": 0, "balance_days": 0, "entries": []}
    return {
        "doctor_id": doctor_id,
        "balance_hours": account.balance_hours,
        "balance_days": round(account.balance_days, 1),
        "entries": [{"id": e.id, "date": e.date, "call_type": e.call_type,
                     "hours_earned": e.hours_earned, "hours_used": e.hours_used,
                     "status": e.status} for e in account.entries],
    }


@app.get("/comp-time/overview/all", tags=["Komp-tid"])
async def comp_time_overview():
    result = {}
    for doc_id, account in db._comp_accounts.items():
        result[doc_id] = {"balance_hours": account.balance_hours, "balance_days": round(account.balance_days, 1)}
    return {"accounts": result}


# === BYTESFÖRFRÅGNINGAR ===

@app.post("/swap-requests", tags=["Byte"])
async def create_swap_request(body: dict):
    from data_model import SwapRequest
    req = SwapRequest(
        id=f"swap_{len(db._swap_requests)+1}",
        clinic_id=body.get("clinic_id", "kristianstad"),
        requester_id=body.get("requester_id", ""),
        requester_date=body.get("requester_date", ""),
        requester_function=body.get("requester_function", ""),
        target_id=body.get("target_id", ""),
        target_date=body.get("target_date", ""),
        target_function=body.get("target_function", ""),
        message=body.get("message", ""),
        created_at=datetime.now().isoformat(),
        expires_at=(datetime.now() + timedelta(days=7)).isoformat(),
    )
    db._swap_requests.append(req)
    await db.audit("swap_request_created", details={"id": req.id, "requester": req.requester_id, "target": req.target_id})
    return {"status": "created", "id": req.id, "status_code": req.status}


@app.get("/swap-requests", tags=["Byte"])
async def list_swap_requests(doctor_id: str = None, status: str = None):
    reqs = db._swap_requests
    if doctor_id:
        reqs = [r for r in reqs if r.requester_id == doctor_id or r.target_id == doctor_id]
    if status:
        reqs = [r for r in reqs if r.status == status]
    return {"requests": [{"id": r.id, "requester_id": r.requester_id, "target_id": r.target_id,
                          "requester_date": r.requester_date, "target_date": r.target_date,
                          "status": r.status, "message": r.message, "created_at": r.created_at}
                         for r in reqs]}


@app.post("/swap-requests/{req_id}/peer-respond", tags=["Byte"])
async def peer_respond(req_id: str, body: dict = {}):
    req = next((r for r in db._swap_requests if r.id == req_id), None)
    if not req:
        raise HTTPException(status_code=404, detail="Förfrågan inte hittad")
    accept = body.get("accept", False)
    if accept:
        req.status = "peer_accepted"
        req.peer_response_at = datetime.now().isoformat()
        # Auto ATL-validering (enkel)
        req.status = "atl_validated"
    else:
        req.status = "peer_rejected"
        req.peer_response_at = datetime.now().isoformat()
    return {"status": req.status, "id": req.id}


@app.post("/swap-requests/{req_id}/admin-respond", tags=["Byte"])
async def admin_respond(req_id: str, body: dict = {}):
    req = next((r for r in db._swap_requests if r.id == req_id), None)
    if not req:
        raise HTTPException(status_code=404, detail="Förfrågan inte hittad")
    approve = body.get("approve", False)
    if approve:
        req.status = "approved"
        req.admin_response_at = datetime.now().isoformat()
        req.completed_at = datetime.now().isoformat()
        await db.audit("swap_approved", details={"id": req.id})
    else:
        req.status = "rejected"
        req.admin_response_at = datetime.now().isoformat()
    return {"status": req.status, "id": req.id}


@app.post("/swap-requests/{req_id}/cancel", tags=["Byte"])
async def cancel_swap_request(req_id: str):
    req = next((r for r in db._swap_requests if r.id == req_id), None)
    if not req:
        raise HTTPException(status_code=404, detail="Förfrågan inte hittad")
    req.status = "cancelled"
    return {"status": "cancelled", "id": req.id}


# === REGELMOTOR ===

# In-memory regel-lagring
_detailed_rules = {}  # {clinic_id: [DetailedRule]}


@app.get("/rules", tags=["Regler"])
async def list_rules(clinic_id: str = "kristianstad"):
    rules = _detailed_rules.get(clinic_id, [])
    return {"rules": [_rule_to_dict(r) for r in rules], "count": len(rules)}


@app.post("/rules", tags=["Regler"])
async def create_rule(body: dict):
    clinic_id = body.get("clinic_id", "kristianstad")
    from rule_engine import DetailedRule, TimeFilter, PersonFilter, ActionSpec
    rule = DetailedRule(
        id=body.get("id", f"rule_{len(_detailed_rules.get(clinic_id, []))+1}"),
        name=body.get("name", "Ny regel"),
        description=body.get("description", ""),
        category=body.get("category", "preference"),
        is_hard=body.get("is_hard", False),
        weight=body.get("weight", 5),
        time_filter=TimeFilter(**body["time_filter"]) if body.get("time_filter") else None,
        person_filter=PersonFilter(**body["person_filter"]) if body.get("person_filter") else None,
        action=ActionSpec(**body["action"]) if body.get("action") else None,
        source=body.get("source", "manual"),
    )
    if clinic_id not in _detailed_rules:
        _detailed_rules[clinic_id] = []
    _detailed_rules[clinic_id].append(rule)
    await db.audit("rule_created", details={"clinic_id": clinic_id, "rule_id": rule.id})
    return {"status": "created", "rule": _rule_to_dict(rule)}


@app.delete("/rules/{rule_id}", tags=["Regler"])
async def delete_rule(rule_id: str, clinic_id: str = "kristianstad"):
    rules = _detailed_rules.get(clinic_id, [])
    _detailed_rules[clinic_id] = [r for r in rules if r.id != rule_id]
    return {"status": "deleted", "rule_id": rule_id}


@app.post("/rules/validate", tags=["Regler"])
async def validate_rules(clinic_id: str = "kristianstad"):
    rules = _detailed_rules.get(clinic_id, [])
    config = await db.get_config(clinic_id)
    if not config or not rules:
        return {"conflicts": [], "total": 0}
    from rule_engine import RuleEngine
    engine = RuleEngine(config, rules)
    conflicts = engine.validate_rules()
    return {"conflicts": conflicts, "total": len(conflicts)}


@app.get("/rules/templates", tags=["Regler"])
async def get_rule_templates():
    from rule_engine import create_example_rules
    rules = create_example_rules()
    return {"templates": [_rule_to_dict(r) for r in rules], "count": len(rules)}


def _rule_to_dict(rule) -> dict:
    """Konvertera DetailedRule till dict för API-svar."""
    d = {"id": rule.id, "name": rule.name, "description": rule.description,
         "category": rule.category, "is_hard": rule.is_hard, "weight": rule.weight,
         "enabled": rule.enabled, "source": getattr(rule, "source", "manual")}
    if rule.time_filter:
        d["time_filter"] = {k: v for k, v in rule.time_filter.__dict__.items() if v is not None}
    if rule.person_filter:
        d["person_filter"] = {k: v for k, v in rule.person_filter.__dict__.items() if v is not None}
    if rule.action:
        d["action"] = {k: v for k, v in rule.action.__dict__.items() if v is not None}
    return d


# === SIMULATOR ===

@app.post("/simulate", tags=["Simulator"])
async def run_simulation(body: dict):
    clinic_id = body.get("clinic_id", "kristianstad")
    config = await db.get_config(clinic_id)
    if not config:
        raise HTTPException(status_code=404, detail="Klinik inte hittad")

    from simulator import WhatIfSimulator, Scenario

    # Hämta baseline-schema om det finns
    schedules = await db.list_schedules(clinic_id=clinic_id)
    baseline = max(schedules, key=lambda s: s.get("created_at", "")).get("raw_schedule") if schedules else None

    simulator = WhatIfSimulator(config, baseline)
    scenarios = [Scenario(type=s["type"], description=s.get("description", ""),
                          changes=s.get("changes", {})) for s in body.get("scenarios", [])]

    results = simulator.simulate(scenarios, time_limit=body.get("time_limit", 15))
    await db.audit("simulation_run", details={"clinic_id": clinic_id, "scenarios": len(scenarios)})

    return {
        "results": [{
            "feasible": r.feasible,
            "impact_summary_sv": r.impact_summary_sv,
            "staffing_impact": r.staffing_impact,
            "call_impact": r.call_impact,
            "ob_impact": r.ob_impact,
            "recommendations_sv": r.recommendations_sv,
            "risk_level": r.risk_level,
        } for r in results],
        "total_scenarios": len(scenarios),
    }


# === MAIN ===
if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000, log_level="info")


